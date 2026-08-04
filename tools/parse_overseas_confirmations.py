"""해외거래 체결확인서(Outlook) 파싱 → .cache/pending_trades.json

배경 (2026-08-03):
해외 ETF 는 체결 다음 영업일(브로커 SettleDate)에야 원장(dt.DWPM10520)에 잡힌다.
그래서 체결 당일 저녁~다음날 아침 사이엔 대시보드가 매도/매수를 전혀 모른다.
실측: 2JM23 IAUM 매도(체결 7/31, Net USD 765,198.83)가 8/3 결제일까지 원장에 없음.

확인서 메일은 체결 다음날 새벽에 도착하므로, 이 배치가 메일을 먼저 읽어
"결제 예정" 목록을 만들어 두면 원장이 따라잡기 전까지 화면에 안내할 수 있다.

동작:
  Outlook COM → PST '업무\\해외거래' 폴더 → xls/xlsx 첨부 파싱 → JSON 저장.
  API(holdings_service)는 이 JSON 만 읽으므로 Outlook 실행 여부와 무관하게 동작한다.

실행:
  python -m tools.parse_overseas_confirmations
  python -m tools.parse_overseas_confirmations --days 90 --verbose

확인서 양식 (한국투자증권·DB증권 공통 17컬럼, 2026-08-03 실측):
  Order No / Fund code / Broker code / TICKER / ISIN / Security Name / B/S / CCY /
  OrderDate / SettleDate / Qty / Price / Gross / Commission / Tax / Fee / Net Amount
  - 헤더 대소문자가 브로커마다 다름 ('Fund code' vs 'Fund Code') → 소문자로 정규화
  - 날짜 포맷도 다름 (KIS '2026-07-31' vs DB증권 '20260720') → _norm_date 로 흡수
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from datetime import datetime, timedelta

# 폴더 경로는 FolderPath 접미사로 매칭 (스토어명 변동 대비)
FOLDER_SUFFIX = r"\업무\해외거래"

_OUT_DEFAULT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    ".cache", "pending_trades.json",
)

# 확인서 컬럼 → 내부 키 (헤더 소문자 기준)
_COLS = {
    "fund code": "fund_cd",
    "ticker": "ticker",
    "isin": "isin",
    "security name": "security_name",
    "b/s": "side",
    "ccy": "ccy",
    "orderdate": "order_date",
    "settledate": "settle_date",
    "qty": "qty",
    "price": "price",
    "gross": "gross",
    "net amount": "net_amount",
}

# DRM 래핑 판별 (사내 생성 파일). 외부 수신 첨부는 보통 클린이지만 방어적으로 확인.
_DRM_MAGIC = b"<DOCUMENT SAFER"


def _norm_date(v) -> str | None:
    """'2026-07-31' / '20260720' / Excel serial → 'YYYY-MM-DD'."""
    if v is None or v == "":
        return None
    if isinstance(v, float) and v > 20000:      # Excel serial (1900 기준)
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(v))).strftime("%Y-%m-%d")
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"
    return None


def _num(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def _read_rows(path: str) -> list[dict]:
    """확인서 1개 파일 → row dict 리스트. 양식이 아니면 빈 리스트."""
    with open(path, "rb") as fh:
        head = fh.read(16)
    if head.startswith(_DRM_MAGIC):
        print(f"  [skip] DRM 래핑 파일: {os.path.basename(path)}", file=sys.stderr)
        return []

    if path.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        sh = wb[wb.sheetnames[0]]
        grid = [list(r) for r in sh.iter_rows(values_only=True)]
        wb.close()
    else:
        import xlrd
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        grid = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]

    if len(grid) < 2:
        return []

    hdr = [str(h or "").strip().lower() for h in grid[0]]
    if "isin" not in hdr or "settledate" not in hdr:
        return []                                   # 확인서 양식 아님

    idx = {_COLS[h]: i for i, h in enumerate(hdr) if h in _COLS}
    rows = []
    for raw in grid[1:]:
        def g(key):
            i = idx.get(key)
            return raw[i] if i is not None and i < len(raw) else None

        fund_cd = str(g("fund_cd") or "").strip()
        settle = _norm_date(g("settle_date"))
        if not fund_cd or not settle:
            continue
        rows.append({
            "fund_cd": fund_cd,
            "ticker": str(g("ticker") or "").strip(),
            "isin": str(g("isin") or "").strip(),
            "security_name": str(g("security_name") or "").strip(),
            "side": str(g("side") or "").strip().upper(),   # B / S
            "ccy": str(g("ccy") or "").strip().upper(),
            "order_date": _norm_date(g("order_date")),
            "settle_date": settle,
            "qty": _num(g("qty")),
            "price": _num(g("price")),
            "gross": _num(g("gross")),
            "net_amount": _num(g("net_amount")),
        })
    return rows


def _outlook_folder():
    """PST '업무\\해외거래' 폴더 반환. 못 찾으면 RuntimeError."""
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    try:
        app = win32com.client.GetActiveObject("Outlook.Application")
    except Exception:
        app = win32com.client.Dispatch("Outlook.Application")
    ns = app.GetNamespace("MAPI")

    def walk(folder, depth):
        if depth > 5:
            return None
        try:
            if str(folder.FolderPath).endswith(FOLDER_SUFFIX):
                return folder
        except Exception:
            return None
        subs = folder.Folders
        for i in range(1, subs.Count + 1):
            hit = walk(subs.Item(i), depth + 1)
            if hit is not None:
                return hit
        return None

    for s in range(1, ns.Folders.Count + 1):
        hit = walk(ns.Folders.Item(s), 0)
        if hit is not None:
            return hit
    raise RuntimeError(f"Outlook 폴더 미발견: ...{FOLDER_SUFFIX}")


def collect(days: int = 60, verbose: bool = False) -> list[dict]:
    """해외거래 폴더의 최근 확인서를 파싱해 거래 리스트 반환 (중복 제거됨)."""
    folder = _outlook_folder()
    items = folder.Items
    items.Sort("[ReceivedTime]", True)
    cutoff = datetime.now() - timedelta(days=days)

    tmpdir = tempfile.mkdtemp(prefix="ocio_conf_")
    trades: list[dict] = []
    seen: set[tuple] = set()

    for i in range(1, items.Count + 1):
        m = items.Item(i)
        try:
            received = datetime(*m.ReceivedTime.timetuple()[:6])
        except Exception:
            continue
        if received < cutoff:
            break                                   # 정렬돼 있으므로 이후는 전부 오래됨

        subject = str(getattr(m, "Subject", ""))
        sender = str(getattr(m, "SenderName", ""))
        atts = m.Attachments
        for a in range(1, atts.Count + 1):
            att = atts.Item(a)
            fn = str(att.FileName)
            if not re.search(r"\.(xls|xlsx)$", fn, re.I):
                continue
            # 같은 메일에 펀드별 첨부가 여러 개 → 파일명 충돌 방지 (i/a 로 유니크)
            path = os.path.join(tmpdir, f"m{i}_a{a}{os.path.splitext(fn)[1]}")
            try:
                att.SaveAsFile(path)
                rows = _read_rows(path)
            except Exception as exc:               # noqa: BLE001 — 1건 실패로 전체 중단 금지
                print(f"  [warn] {fn}: {type(exc).__name__} {exc}", file=sys.stderr)
                continue
            for r in rows:
                key = (r["fund_cd"], r["isin"], r["order_date"], r["side"],
                       r["qty"], r["net_amount"])
                if key in seen:                     # 재발송/중복 메일
                    continue
                seen.add(key)
                r["broker"] = sender
                r["mail_subject"] = subject
                r["mail_received"] = received.strftime("%Y-%m-%d %H:%M")
                r["source_file"] = fn
                trades.append(r)
            if verbose:
                print(f"  {received:%Y-%m-%d} {fn} → {len(rows)}행")

    trades.sort(key=lambda r: (r["settle_date"], r["fund_cd"]), reverse=True)
    return trades


def main() -> int:
    ap = argparse.ArgumentParser(description="해외거래 확인서 파싱 → pending_trades.json")
    ap.add_argument("--days", type=int, default=60, help="조회 기간 (기본 60일)")
    ap.add_argument("--out", default=_OUT_DEFAULT, help="출력 JSON 경로")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    try:
        trades = collect(days=args.days, verbose=args.verbose)
    except Exception as exc:                        # noqa: BLE001 — Outlook 미실행 등
        print(f"[FAIL] {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "days": args.days,
        "trades": trades,
    }
    # 원자적 교체 — API 가 읽는 중에 부분 파일을 보지 않도록
    tmp = args.out + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    os.replace(tmp, args.out)

    funds = sorted({t["fund_cd"] for t in trades})
    print(f"[OK] {len(trades)}건 / 펀드 {len(funds)}개 ({', '.join(funds)}) → {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
