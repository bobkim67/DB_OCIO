# -*- coding: utf-8 -*-
"""DB생명(4JM12) 월간 운용보고 데이터 → 엑셀 (PPT 페이지별 시트, 2026-07-31 사용자 지시).

발송 PPT((YYYY-MM)DB생명 글로벌Active자산배분혼합형.ppt) 6페이지의 데이터를
DB 값으로 재현해 시트별로 쓴다. 서식은 깔끔한 표 기준(PPT 디자인 복제 아님).

  s2 펀드개요·NAV일별·기간수익률 : dt.DWPM10510 + DT BM(DWPM10040 'B')
  s3 보유비중·자산구성·환헤지    : dt.DWPM10530 월말 (환헤지 산식은 아래 주석)
  s4 수익증권 보유 현황          : dt.DWPM10530 (원화 + 외화 FC_*)
  s5 매매내역                    : dt.DWPM10520 월간 합산
  s6 운용경과·운용계획           : report_output/{YYYY-MM}/4JM12.final.json (있으면)

실행:  python -m tools.dblife_monthly_excel --month 2026-06
출력:  output/DB생명_월간보고_데이터_{YYYYMM}.xlsx (기본)

환헤지 표 정의 (발송 PPT 기준, memory db-life-monthly-report):
  헤지 포지션/순자산 = 미국달러선물(매도) 비중(양수)
  해외자산/순자산   = 미국주식 ETF 비중 합(+USD 예치금)
  BM 기준           = 헤지 / 45%(BM 해외비중)
  USD Exposure 기준 = 헤지 / 해외자산비중
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from dateutil.relativedelta import relativedelta

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

from modules.data_loader import get_pandas_connection, load_dt_bm_prices  # noqa: E402

FUND = '4JM12'
FUND_NM = '(무)DB변액글로벌Active자산배분혼합형'
INCEPTION = '2022-03-18'
INCEPTION_BASE = 1970.76      # 설정후 분모 (reference_4jm12_inception_base)
BM_BASE = 1000.0              # DT BM 설정후 분모
MANAGER = '강성수'            # 담당운용역 (PPT 고정 표기)
BM_NOTE = ('※ BM : (0.55 * [KBP-동부생명7 총수익지수]) + (0.225 * [MSCI ACWI Standard '
           'Index_U(USD)(t-1)]) + (0.225 * [MSCI ACWI Standard Index_U(USDKRW)(t-1)(당일환율)])')

# 자산구성(파이) 분류: 자산분류코드명 기반 — 채권형/주식형, 나머지는 유동성 잔여
_BOND_KW = ('채권',)
_EQUITY_KW = ('주식', '지수')
_FX_FUT_KW = ('달러', '통화선물', 'NDF')


def _ym_range(month: str) -> tuple[str, str, str]:
    """'2026-06' → (월초 '20260601', 월말 '20260630', 표시 '2026.06.30')."""
    y, m = int(month[:4]), int(month[5:7])
    last = (date(y, m, 1) + relativedelta(months=1) - timedelta(days=1))
    return f'{y:04d}{m:02d}01', last.strftime('%Y%m%d'), last.strftime('%Y.%m.%d')


def _q(sql: str, params: list, db: str = 'dt') -> pd.DataFrame:
    conn = get_pandas_connection(db)
    try:
        return pd.read_sql(sql, conn, params=params)
    finally:
        conn.close()


# ── 데이터 로드 ───────────────────────────────────────────────────


def load_nav_daily(start: str, end: str) -> pd.DataFrame:
    """월간 일별 NAV/기준가 (원장은 캘린더일 행 보유 — 주말 보수 accrual 포함)."""
    df = _q(
        "SELECT STD_DT, NAST_AMT, MOD_STPR, OCPY_AMT FROM DWPM10510 "
        "WHERE IMC_CD='003228' AND FUND_CD=%s AND STD_DT BETWEEN %s AND %s "
        "ORDER BY STD_DT", [FUND, str(start), str(end)])
    df['일자'] = pd.to_datetime(df['STD_DT'].astype(str))
    return df


def load_stpr_series() -> pd.Series:
    df = _q("SELECT STD_DT, MOD_STPR FROM DWPM10510 "
            "WHERE IMC_CD='003228' AND FUND_CD=%s ORDER BY STD_DT", [FUND])
    s = pd.Series(df['MOD_STPR'].values,
                  index=pd.to_datetime(df['STD_DT'].astype(str))).astype(float)
    return s[~s.index.duplicated(keep='last')]


def load_bm_series() -> pd.Series:
    d = load_dt_bm_prices(FUND, '20200101')
    s = d.sort_values('기준일자').set_index('기준일자')['value'].astype(float)
    return s[~s.index.duplicated(keep='last')]


def period_returns(s: pd.Series, end_dt: pd.Timestamp, base: float) -> dict:
    """PPT 기간수익률 — 앵커 = **N개월 전 월말** (2026-06 발송본 역산으로 확정).

    relativedelta(end-3M=3/30)가 아니라 월말일(3/31)이 분모다: 3M 15.83%(=2분기
    수익률)·6M 8.29% 는 3/31·12/31 앵커로만 재현된다. 1M=전월말, 1Y=12개월전 월말.
    """
    def at(dt_):
        sub = s[s.index <= dt_]
        return float(sub.iloc[-1]) if len(sub) else None

    def month_end_before(n_months: int) -> pd.Timestamp:
        return (end_dt.replace(day=1) - relativedelta(months=n_months - 1)
                - timedelta(days=1))
    endv = at(end_dt)
    out = {}
    for label, n in (('최근1개월', 1), ('최근3개월', 3), ('최근6개월', 6), ('최근1년', 12)):
        v0 = at(month_end_before(n))
        out[label] = (endv / v0 - 1) * 100 if (endv and v0) else None
    out['설정일 이후'] = (endv / base - 1) * 100 if endv else None
    return out


def load_holdings_eom(eom: str) -> pd.DataFrame:
    df = _q(
        "SELECT ITEM_CD, ITEM_NM, AST_CLSF_CD_NM, CURR_DS_CD, POS_DS_CD, "
        "       QTY, ACQ_AMT, EVL_AMT, EVL_PL_AMT, NAST_TAMT_AGNST_WGH, "
        "       FC_ACQ_AMT, FC_EVL_AMT, FC_PL_AMT "
        "FROM DWPM10530 "
        "WHERE IMC_CD='003228' AND FUND_CD=%s AND STD_DT=%s "
        "  AND ITEM_NM NOT LIKE '%%미지급%%' AND ITEM_NM NOT LIKE '%%미수%%' "
        "ORDER BY EVL_AMT DESC", [FUND, str(eom)])
    for c in ('QTY', 'ACQ_AMT', 'EVL_AMT', 'EVL_PL_AMT', 'NAST_TAMT_AGNST_WGH',
              'FC_ACQ_AMT', 'FC_EVL_AMT', 'FC_PL_AMT'):
        df[c] = pd.to_numeric(df[c], errors='coerce')
    return df


def load_trades_month(start: str, end: str) -> pd.DataFrame:
    df = _q(
        "SELECT t.ITEM_CD, t.ITEM_NM, t.AST_CLSF_CD, t.BUY_SELL_DS_CD, t.TR_CD, "
        "       t.CURR_DS_CD, t.TRD_QTY, t.TRD_AMT, t.STL_AMT, t.KRW_STL_AMT, "
        "       t.TRD_PL_AMT, t.KRW_TRD_PL_AMT, c.tr_nm AS TR_NM "
        "FROM DWPM10520 t "
        "LEFT JOIN DWCI10160 c ON t.tr_cd = c.tr_cd AND t.synp_cd = c.synp_cd "
        "WHERE t.IMC_CD='003228' AND t.FUND_CD=%s AND t.STD_DT BETWEEN %s AND %s",
        [FUND, str(start), str(end)])
    for c in ('TRD_QTY', 'TRD_AMT', 'STL_AMT', 'KRW_STL_AMT', 'TRD_PL_AMT', 'KRW_TRD_PL_AMT'):
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)
    return df


def load_final_comment(month: str) -> str:
    p = BASE / 'market_research' / 'data' / 'report_output' / month / f'{FUND}.final.json'
    if not p.exists():
        return ''
    try:
        d = json.loads(p.read_text(encoding='utf-8'))
        return str(d.get('final_comment') or '')
    except Exception:
        return ''


# ── 분류 헬퍼 ─────────────────────────────────────────────────────


def bucket_of(row) -> str:
    """채권형/주식형/달러선물/외화현금/유동성 — s3·s5 구분용."""
    nm = str(row['ITEM_NM'] or '')
    ast = str(row['AST_CLSF_CD_NM'] or '') if 'AST_CLSF_CD_NM' in row else ''
    if any(k in nm for k in ('미국달러', '달러 F', '통화선물')) or '달러선물' in ast:
        return '달러선물'
    if 'USD' in nm.upper() and ('DEPOSIT' in nm.upper() or '예치' in nm):
        return '외화현금'
    if any(k in ast for k in _BOND_KW) or any(k in nm for k in ('국고채', '채권', 'TMF')):
        return '채권형'
    if any(k in ast for k in _EQUITY_KW) or row.get('CURR_DS_CD') not in (None, '', 'KRW'):
        return '주식형'
    return '유동성'


# ── 엑셀 쓰기 ─────────────────────────────────────────────────────


def build(month: str, out_path: Path) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    start, eom, eom_disp = _ym_range(month)
    end_dt = pd.Timestamp(eom)

    nav = load_nav_daily(start, eom)
    stpr = load_stpr_series()
    bm = load_bm_series()
    hold = load_holdings_eom(eom)
    trades = load_trades_month(start, eom)
    comment = load_final_comment(month)

    if nav.empty or hold.empty:
        raise SystemExit(f'{month} 데이터 없음 (NAV {len(nav)}행 / 보유 {len(hold)}행)')

    ap_ret = period_returns(stpr, end_dt, INCEPTION_BASE)
    bm_ret = period_returns(bm, end_dt, BM_BASE)

    # ── 스타일 ──
    H_FILL = PatternFill('solid', fgColor='DCE6F1')
    T_FILL = PatternFill('solid', fgColor='4A3728')
    BOLD = Font(bold=True)
    TITLE = Font(bold=True, size=13, color='FFFFFF')
    SEC = Font(bold=True, size=11, color='1F3864')
    thin = Side(style='thin', color='B7B7B7')
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
    R = Alignment(horizontal='right')
    C = Alignment(horizontal='center')

    wb = Workbook()

    def sheet(name):
        ws = wb.create_sheet(name) if wb.sheetnames != ['Sheet'] else wb.active
        ws.title = name
        return ws

    def put(ws, r, c, v, *, bold=False, fill=None, fmt=None, align=None, border=True):
        cell = ws.cell(row=r, column=c, value=v)
        if bold:
            cell.font = BOLD
        if fill is not None:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        if align:
            cell.alignment = align
        if border:
            cell.border = BOX
        return cell

    def title(ws, r, text, ncols):
        ws.cell(row=r, column=1, value=text).font = TITLE
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = T_FILL
        return r + 2

    def section(ws, r, text):
        ws.cell(row=r, column=1, value=text).font = SEC
        return r + 1

    NUM = '#,##0'
    NUM2 = '#,##0.00'
    PCT2 = '0.00'

    # ── s1 표지 ──
    ws = sheet('1.표지')
    r = title(ws, 1, f'DB생명 월간 운용보고 — {month}', 6)
    for label, val in (('펀드명', FUND_NM), ('펀드코드', FUND), ('기준일', eom_disp),
                       ('운용사', '한국투자신탁운용'), ('담당운용역', MANAGER),
                       ('생성', f'DB 재현 (dt 원장), {datetime.now():%Y-%m-%d}')):
        put(ws, r, 1, label, bold=True, fill=H_FILL)
        put(ws, r, 2, val)
        r += 1
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 44

    # ── s2 펀드개요 + NAV 일별 + 기간수익률 ──
    ws = sheet('2.개요_NAV_수익률')
    last = nav.iloc[-1]
    r = title(ws, 1, f'I. 운용 현황 — 기준일 {eom_disp}', 8)
    r = section(ws, r, '펀드 개요')
    for label, val, fmt in (('펀드명', FUND_NM, None),
                            ('설정일', INCEPTION, None),
                            ('설정액(억원)', float(last['OCPY_AMT']) / 1e8, NUM2),
                            ('순자산(억원)', float(last['NAST_AMT']) / 1e8, NUM2),
                            ('기준가(원)', float(last['MOD_STPR']), NUM2),
                            ('담당운용역', MANAGER, None)):
        put(ws, r, 1, label, bold=True, fill=H_FILL)
        put(ws, r, 2, val, fmt=fmt, align=R if fmt else None)
        r += 1

    r += 1
    r = section(ws, r, '펀드 NAV 일별 추이  (단위: 억원, 원)')
    for c, h in enumerate(('기준일', 'NAV(억원)', '기준가(원)'), start=1):
        put(ws, r, c, h, bold=True, fill=H_FILL, align=C)
    r += 1
    for _, row in nav.iterrows():
        put(ws, r, 1, row['일자'].strftime('%Y-%m-%d'), align=C)
        put(ws, r, 2, float(row['NAST_AMT']) / 1e8, fmt=NUM2, align=R)
        put(ws, r, 3, float(row['MOD_STPR']), fmt=NUM2, align=R)
        r += 1

    r += 1
    r = section(ws, r, '기간수익률  (단위: %, %p)')
    cols = ('구분', '최근1개월', '최근3개월', '최근6개월', '최근1년', '설정일 이후')
    for c, h in enumerate(cols, start=1):
        put(ws, r, c, h, bold=True, fill=H_FILL, align=C)
    r += 1
    for label, d in (('기간수익률', ap_ret), ('BM수익률', bm_ret)):
        put(ws, r, 1, label, bold=True, fill=H_FILL)
        for c, k in enumerate(cols[1:], start=2):
            put(ws, r, c, round(d[k], 2) if d[k] is not None else None, fmt=PCT2, align=R)
        r += 1
    put(ws, r, 1, '초과수익률', bold=True, fill=H_FILL)
    for c, k in enumerate(cols[1:], start=2):
        v = (ap_ret[k] - bm_ret[k]) if (ap_ret[k] is not None and bm_ret[k] is not None) else None
        put(ws, r, c, round(v, 2) if v is not None else None, fmt=PCT2, align=R)
    r += 2
    ws.cell(row=r, column=1, value=BM_NOTE)
    for col, w in (('A', 16), ('B', 14), ('C', 14), ('D', 14), ('E', 14), ('F', 14)):
        ws.column_dimensions[col].width = w

    # ── s3 보유비중 + 자산구성 + 환헤지 ──
    ws = sheet('3.투자현황')
    h = hold.copy()
    h['bucket'] = h.apply(bucket_of, axis=1)
    h['w'] = h['NAST_TAMT_AGNST_WGH'].astype(float)

    r = title(ws, 1, '업종별·지역별 투자 현황', 6)
    r = section(ws, r, '펀드 주요 보유 현황')
    for c, hd in enumerate(('구분', '종목', '비중(%)'), start=1):
        put(ws, r, c, hd, bold=True, fill=H_FILL, align=C)
    r += 1
    # PPT 표기 규약: 소수 1자리 **버림**(truncate) — 38.949→38.9, 합계도 원시합 버림
    # (47.986→47.9). 반올림이면 48.0/46.0 이 되어 발송본(47.9/45.9)과 어긋난다.
    trunc1 = lambda x: int(float(x) * 10) / 10
    comp = {}
    for bucket in ('채권형', '주식형'):
        sub = h[h['bucket'] == bucket].sort_values('w', ascending=False)
        wsum = trunc1(sub['w'].sum())
        comp[bucket] = wsum
        put(ws, r, 1, bucket, bold=True, fill=H_FILL)
        put(ws, r, 2, '합계', bold=True)
        put(ws, r, 3, wsum, fmt='0.0', align=R, bold=True)
        r += 1
        for _, row in sub.iterrows():
            put(ws, r, 1, '')
            put(ws, r, 2, f"  {row['ITEM_NM']}")
            put(ws, r, 3, trunc1(row['w']), fmt='0.0', align=R)
            r += 1
    liq = round(100.0 - comp.get('채권형', 0) - comp.get('주식형', 0), 1)

    r += 1
    r = section(ws, r, '펀드 자산 구성 비중 (%)')
    for c, hd in enumerate(('채권형', '주식형', '유동성'), start=1):
        put(ws, r, c, hd, bold=True, fill=H_FILL, align=C)
    r += 1
    put(ws, r, 1, round(comp.get('채권형', 0), 1), fmt='0.0', align=R)
    put(ws, r, 2, round(comp.get('주식형', 0), 1), fmt='0.0', align=R)
    put(ws, r, 3, round(liq, 1), fmt='0.0', align=R)
    r += 2

    # 환헤지 — 헤지 = 달러선물(매도) 비중 절대값, 해외자산 = 주식형(미국 ETF) + 외화현금
    hedge_w = float(h.loc[h['bucket'] == '달러선물', 'w'].abs().sum())
    ovs_w = float(h.loc[h['bucket'] == '주식형', 'w'].sum()
                  + h.loc[h['bucket'] == '외화현금', 'w'].sum())
    r = section(ws, r, '해외자산 환헤지 비율')
    put(ws, r, 1, 'BM 기준(헤지/45%)', bold=True, fill=H_FILL)
    put(ws, r, 2, '=', border=False)
    put(ws, r, 3, round(hedge_w / 45.0 * 100, 1), fmt='0.0', align=R)
    r += 1
    put(ws, r, 1, 'USD Exposure 기준(헤지/해외자산)', bold=True, fill=H_FILL)
    put(ws, r, 2, '=', border=False)
    put(ws, r, 3, round(hedge_w / ovs_w * 100, 1) if ovs_w else None, fmt='0.0', align=R)
    r += 2
    r = section(ws, r, '환헤지 포지션 | 해외자산 비율')
    put(ws, r, 1, '헤지 포지션/순자산(%)', bold=True, fill=H_FILL)
    put(ws, r, 2, round(hedge_w, 1), fmt='0.0', align=R)
    r += 1
    put(ws, r, 1, '해외자산/순자산(%)', bold=True, fill=H_FILL)
    put(ws, r, 2, round(ovs_w, 1), fmt='0.0', align=R)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 12

    # ── s4 수익증권 보유 현황 ──
    ws = sheet('4.보유현황')
    r = title(ws, 1, '수익증권 보유 현황', 8)
    sec_rows = h[h['bucket'].isin(('채권형', '주식형'))]
    krw_rows = sec_rows[(sec_rows['CURR_DS_CD'].isna()) | (sec_rows['CURR_DS_CD'].isin(('', 'KRW')))]
    fc_rows = sec_rows[~sec_rows.index.isin(krw_rows.index)]

    r = section(ws, r, '수익증권 보유 현황  (단위: 원, %)')
    for c, hd in enumerate(('종목명', '수량', '취득단가', '취득액', '평가액', '평가손익', '손익률(%)'), start=1):
        put(ws, r, c, hd, bold=True, fill=H_FILL, align=C)
    r += 1
    for _, row in krw_rows.iterrows():
        qty, acq = float(row['QTY'] or 0), float(row['ACQ_AMT'] or 0)
        # 취득단가: 좌수형(수익증권)은 1,000좌당 — ETF/주식은 1주당
        per = (acq / qty * 1000) if ('투자신탁' in str(row['ITEM_NM'])) else (acq / qty) if qty else 0
        put(ws, r, 1, row['ITEM_NM'])
        put(ws, r, 2, qty, fmt=NUM, align=R)
        put(ws, r, 3, round(per, 2), fmt=NUM2, align=R)
        put(ws, r, 4, acq, fmt=NUM, align=R)
        put(ws, r, 5, float(row['EVL_AMT'] or 0), fmt=NUM, align=R)
        put(ws, r, 6, float(row['EVL_PL_AMT'] or 0), fmt=NUM, align=R)
        put(ws, r, 7, round(float(row['EVL_PL_AMT'] or 0) / acq * 100, 2) if acq else None,
            fmt=PCT2, align=R)
        r += 1

    r += 1
    r = section(ws, r, '외화수익증권 보유 현황  (단위: USD, 원)')
    for c, hd in enumerate(('종목명', '수량', '외화취득액', '외화평가액', '외화평가손익',
                            '원화취득액', '원화평가액', '원화평가손익'), start=1):
        put(ws, r, c, hd, bold=True, fill=H_FILL, align=C)
    r += 1
    tot = [0.0] * 6
    for _, row in fc_rows.iterrows():
        vals = [float(row['FC_ACQ_AMT'] or 0), float(row['FC_EVL_AMT'] or 0),
                float(row['FC_PL_AMT'] or 0), float(row['ACQ_AMT'] or 0),
                float(row['EVL_AMT'] or 0), float(row['EVL_PL_AMT'] or 0)]
        put(ws, r, 1, row['ITEM_NM'])
        put(ws, r, 2, float(row['QTY'] or 0), fmt=NUM2, align=R)
        for c, v in enumerate(vals, start=3):
            put(ws, r, c, v, fmt=NUM2 if c <= 5 else NUM, align=R)
        tot = [a + b for a, b in zip(tot, vals)]
        r += 1
    put(ws, r, 1, '총계', bold=True, fill=H_FILL)
    put(ws, r, 2, None)
    for c, v in enumerate(tot, start=3):
        put(ws, r, c, v, fmt=NUM2 if c <= 5 else NUM, align=R, bold=True)
    ws.column_dimensions['A'].width = 40
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ── s5 매매내역 ──
    ws = sheet('5.매매내역')
    r = title(ws, 1, f'매매내역  (매매기간: {start[:4]}.{int(start[4:6])}.{int(start[6:8])}'
                     f'~{eom[:4]}.{int(eom[4:6])}.{int(eom[6:8])})', 10)
    # ── PPT 규약 (2026-06 발송본 역산으로 확정) ──
    #  · KR상장 증권: 매수/매도 = **결제금액(STL_AMT)** 합 (수수료·세금 포함), 손익 = TRD_PL_AMT
    #  · 외화수익증권: 원화 = KRW_STL_AMT / 원화손익 = KRW_TRD_PL_AMT,
    #                  외화 = STL_AMT(결제) / 외화손익 = TRD_PL_AMT
    #  · 달러선물: 매수열 = Σ정산손익(TRD_PL_AMT), 매도 0 — 명목금액 표시 안 함
    #  · USD 예금(환전): B650(외화매도원화매입)→매수열 / B652(외화매입원화매도)→매도열,
    #                    KRW_STL_AMT 기준. 외화 매매열은 표시 안 함
    #  · 구분: KR상장(채권ETF 포함)=국내주식 / 통화선물=국내선물 / USD예금=외화현금
    #          / 해외상장=외화수익증권
    t = trades.copy()
    t['is_buy'] = t['BUY_SELL_DS_CD'].astype(str).str.upper().eq('M')
    t['is_sell'] = t['BUY_SELL_DS_CD'].astype(str).str.upper().eq('D')
    t['is_fc'] = ~t['CURR_DS_CD'].isna() & ~t['CURR_DS_CD'].isin(('', 'KRW'))

    def _gubun(nm: str, any_fc: bool) -> str:
        u = nm.upper()
        if '달러 F' in nm or '통화선물' in nm or ('미국달러' in nm and ' F ' in f' {nm} '):
            return '국내선물'
        if 'USD' in u and ('DEPOSIT' in u or '예치' in nm):
            return '외화현금'
        return '외화수익증권' if any_fc else '국내주식'

    rows5 = []
    for (cd, nm), x in t.groupby(['ITEM_CD', 'ITEM_NM']):
        gb = _gubun(str(nm), bool(x['is_fc'].any()))
        if gb == '국내선물':
            pl = float(x['TRD_PL_AMT'].sum())
            rec = dict(매수원화=pl, 매도원화=0.0, 손익원화=pl,
                       매수외화=0.0, 매도외화=0.0, 손익외화=0.0)
        elif gb == '외화현금':
            rec = dict(
                매수원화=float(x.loc[x['TR_CD'] == 'B650', 'KRW_STL_AMT'].sum()),
                매도원화=float(x.loc[x['TR_CD'] == 'B652', 'KRW_STL_AMT'].sum()),
                손익원화=float(x['KRW_TRD_PL_AMT'].sum()),
                매수외화=0.0, 매도외화=0.0, 손익외화=0.0)
        elif gb == '외화수익증권':
            rec = dict(
                매수원화=float(x.loc[x['is_buy'], 'KRW_STL_AMT'].sum()),
                매도원화=float(x.loc[x['is_sell'], 'KRW_STL_AMT'].sum()),
                손익원화=float(x['KRW_TRD_PL_AMT'].sum()),
                매수외화=float(x.loc[x['is_buy'], 'STL_AMT'].sum()),
                매도외화=float(x.loc[x['is_sell'], 'STL_AMT'].sum()),
                손익외화=float(x['TRD_PL_AMT'].sum()))
        else:   # 국내주식 (KR상장 증권)
            rec = dict(
                매수원화=float(x.loc[x['is_buy'], 'STL_AMT'].sum()),
                매도원화=float(x.loc[x['is_sell'], 'STL_AMT'].sum()),
                손익원화=float(x.loc[x['is_sell'], 'TRD_PL_AMT'].sum()),
                매수외화=0.0, 매도외화=0.0, 손익외화=0.0)
        rows5.append({'ITEM_CD': cd, 'ITEM_NM': nm, 'bucket': gb, **rec})
    g = pd.DataFrame(rows5)

    # 월말 보유수량/비중 병합
    eom_pos = h.set_index('ITEM_CD')[['QTY', 'w']]
    g = g.join(eom_pos, on='ITEM_CD')
    g['QTY'] = g['QTY'].fillna(0.0)
    g['w'] = g['w'].fillna(0.0)
    _gb_order = {'국내주식': 0, '국내선물': 1, '외화현금': 2, '외화수익증권': 3}
    g['_o'] = g['bucket'].map(_gb_order).fillna(9)
    g = g.sort_values(['_o', 'ITEM_NM']).drop(columns='_o')

    for c, hd in enumerate(('구분', '종목명', '보유수량', '비중(%)', '매수(원화)', '매도(원화)',
                            '매매손익(원화)', '매수(외화)', '매도(외화)', '매매손익(외화)'), start=1):
        put(ws, r, c, hd, bold=True, fill=H_FILL, align=C)
    r += 1
    tots = [0.0] * 3
    for _, row in g.iterrows():
        put(ws, r, 1, row['bucket'], align=C)
        put(ws, r, 2, row['ITEM_NM'])
        put(ws, r, 3, float(row['QTY']), fmt=NUM, align=R)
        put(ws, r, 4, round(float(row['w']), 2), fmt=PCT2, align=R)
        for c, k in enumerate(('매수원화', '매도원화', '손익원화'), start=5):
            put(ws, r, c, float(row[k]), fmt=NUM, align=R)
        for c, k in enumerate(('매수외화', '매도외화', '손익외화'), start=8):
            put(ws, r, c, float(row[k]), fmt=NUM2, align=R)
        tots[0] += row['매수원화']
        tots[1] += row['매도원화']
        tots[2] += row['손익원화']
        r += 1
    put(ws, r, 2, '합계', bold=True, fill=H_FILL)
    for c, v in enumerate(tots, start=5):
        put(ws, r, c, v, fmt=NUM, align=R, bold=True)
    ws.column_dimensions['B'].width = 36
    for i in (3, 5, 6, 7, 8, 9, 10):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ── s6 운용경과·운용계획 ──
    ws = sheet('6.운용경과_계획')
    r = title(ws, 1, 'II. 운용경과 및 운용계획', 4)
    if comment:
        r = section(ws, r, f'승인 코멘트 (report_output/{month}/{FUND}.final.json)')
        for para in comment.split('\n'):
            ws.cell(row=r, column=1, value=para)
            r += 1
    else:
        r = section(ws, r, '승인 코멘트 없음 — 아래 항목 수기 작성')
        for item in ('운용경과 - 시장 동향 :', '운용경과 - 운용 경과 :', '운용경과 - 펀드 성과 :',
                     '운용계획 - 매크로 :', '운용계획 - 운용계획 :', '운용계획 - 환헤지 비율 :'):
            put(ws, r, 1, item, bold=True, fill=H_FILL, border=False)
            r += 2
    ws.column_dimensions['A'].width = 120

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # 검증 출력 — 발송 PPT 와 대조할 핵심 수치
    _s4 = {row['ITEM_NM']: {'수량': float(row['QTY'] or 0), '취득액': float(row['ACQ_AMT'] or 0),
                            '평가액': float(row['EVL_AMT'] or 0), '평가손익': float(row['EVL_PL_AMT'] or 0)}
           for _, row in sec_rows.iterrows()}
    _s5 = {row['ITEM_NM']: {'매수원화': row['매수원화'], '매도원화': row['매도원화'],
                            '손익원화': row['손익원화'], '매수외화': row['매수외화'],
                            '매도외화': row['매도외화'], '보유수량': float(row['QTY']),
                            '비중': round(float(row['w']), 2)}
           for _, row in g.iterrows()}
    return {
        's4_보유': _s4,
        's5_매매': _s5,
        'nav_rows': len(nav),
        '월말 NAV(억)': round(float(last['NAST_AMT']) / 1e8, 2),
        '월말 기준가': round(float(last['MOD_STPR']), 2),
        '기간수익률': {k: round(v, 2) for k, v in ap_ret.items() if v is not None},
        'BM수익률': {k: round(v, 2) for k, v in bm_ret.items() if v is not None},
        '자산구성': {'채권형': round(comp.get('채권형', 0), 1),
                     '주식형': round(comp.get('주식형', 0), 1), '유동성': round(liq, 1)},
        '환헤지': {'헤지/순자산': round(hedge_w, 1), '해외자산/순자산': round(ovs_w, 1),
                   'BM기준': round(hedge_w / 45.0 * 100, 1),
                   'USDExp기준': round(hedge_w / ovs_w * 100, 1) if ovs_w else None},
        'out': str(out_path),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description='DB생명(4JM12) 월간 운용보고 데이터 엑셀 생성')
    ap.add_argument('--month', required=True, help='YYYY-MM (예: 2026-06)')
    ap.add_argument('--out', default=None)
    a = ap.parse_args()
    out = Path(a.out) if a.out else (
        BASE / 'output' / f'DB생명_월간보고_데이터_{a.month.replace("-", "")}.xlsx')
    st = build(a.month, out)
    print(json.dumps(st, ensure_ascii=False, indent=1))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
