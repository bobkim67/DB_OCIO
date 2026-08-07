# -*- coding: utf-8 -*-
"""KB국민은행 투자풀(07G07) 월간 워드 보고 — 붙여넣기용 데이터 엑셀 (2026-08-07 사용자 지시).

★ **Word COM 치환기(`kb_monthly_report.build_docx`)는 쓰지 않는다.** 발송본 워드는
  그대로 열어 두고, 이 엑셀에서 **표 블록을 복사 → 워드 표에 붙여넣기** 한다.
  2026-08-06 에 신한라이프 PPT 치환기를 같은 이유로 폐기했다 — COM 은 서버에
  Office 가 떠야 하고, 사내 DRM 래핑까지 겹쳐 취약하다.
  (`kb_monthly_report.collect()` 는 수치 산출부라 **계속 쓴다**. 폐기 대상은 docx 쓰기부.)

수치 정의는 `kb_monthly_report` 와 **같은 소스**를 쓴다 — 한쪽만 고치면 같은 달
보고서 안에서 값이 갈린다.

## 붙여넣기 규약 — 워드 표와 행·열이 1:1

| 워드 | 격자 | 이 엑셀 |
|---|---|---|
| 표3 성과요인분해 | 9행 × 5열 | `표` 시트 상단 블록 |
| 표4 운용수익률 | 4행 × 7열 (값은 4행 C~G) | 가운데 블록 |
| 표5 투자비중 | 6행 × 10열 (값은 3·6행 C~J) | 하단 2블록 |

각 블록의 **값 영역만** 색을 빼고 그대로 두었다 — 헤더까지 같이 복사해도 워드 표의
헤더와 겹치지 않도록 블록 시작 위치를 표시해 뒀다.

## 확정된 규약

- **표3 헤더 = 기준가 기준** (2026-08-07 사용자 확정). 펀드=07G07 기준가 1개월,
  BM=DT BM1, 초과=차. 이러면 **요인 합계 = 헤더 초과**로 정합한다.
  ⚠ 구 발송본(2026-05)은 헤더에 마스터 PA(+7.39/+2.80/+4.58)를 쓰고 요인 합계는
    기준가 초과(+3.76)라 표 안에서 두 기준이 섞여 있었다.
- **설정 후* = 앵커 SI**(2022-01-03 기준, 2026-08-07 사용자 확정). 2026-07 = +30.15.
  ⚠ 구 발송본은 이 칸에 **설정후 초과**를 넣었다 — 실측으로 확인:
    2026-05 초과 15.4466 = 발송본 +15.44 / 2026-06 초과 11.2953 vs 발송본 +11.26.
    이제는 절대수익률을 쓴다(7월 FactSheet O9=30.1466 과 같은 값).
- **성과요인분해 = FX 분리 · 방법1** (표에 FX 행이 따로 있다). FactSheet 쪽은
  FX 포함 · 방법3 이라 서로 다르다 — 섞지 말 것.
- **유동성 비중 = 잔여(100 − Σ)**. 현금·예금은 손익 소스(MA000410)에 라인이 없어
  PA 비중 합이 100 에 미달한다.

실행:  python -m tools.kb_monthly_excel --month 2026-07
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
FUND = '07G07'
FUND_NAME = '한국투자OCIO 알아서'
COMPANY = '한투운용'
OUT_NAME = 'KB국민은행_07G07_월간워드_데이터_{ym}.xlsx'

# 워드 표3 자산군 행 순서 — 발송본 그대로. 라벨 매칭이 아니라 **행 순서가 곧 정확도**다
# (신한 엑셀에서 겪은 것과 같은 함정 — 순서가 어긋나면 값이 엉뚱한 행에 붙는다).
FACTOR_ROWS = ('주식', '채권', 'FX', '유동성 및 비용')
WEIGHT_KEYS = ('주식', '채권', '대체', '유동성')


@dataclass
class SheetData:
    period: str
    raw: dict
    narrative: dict = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


# ══════════════════════════════════════════════════════════════
# 수집
# ══════════════════════════════════════════════════════════════

def collect(period: str) -> SheetData:
    """수치는 전부 `kb_monthly_report.collect` — 단일 소스."""
    from tools.kb_monthly_report import collect as _collect

    raw = _collect(period)
    d = SheetData(period=period, raw=raw)
    _check(d)
    return d


# ══════════════════════════════════════════════════════════════
# 워드 서술 — 수치는 코드, 문장만 LLM
# ══════════════════════════════════════════════════════════════
#
# 2JM23 성과 문장에서 얻은 교훈을 그대로 적용한다: **순수 수치 나열에 LLM 을 끼우면
# 오기·표현 흔들림만 생긴다.** 헤드라인·환기여·BM대비 초과성과·요인 수치 접두는
# 코드가 쓰고, LLM 은 "왜" 만 쓴다.
#
# 블록 정의 = 2026-05 발송본(월간 포맷) 구조 그대로. 키는 워드 순서.
_BLOCKS: tuple[tuple[str, str, int], ...] = (
    # (키, 워드 위치 라벨, 목표 자수)
    ('절대성과_주식',   '(성과) ㅇ 절대 성과 — 주식',            140),
    ('절대성과_채권',   '(성과) ㅇ 절대 성과 — 채권',            120),
    ('자산배분효과',    '(성과) 자산배분효과(TAA) 해설',          150),
    ('종목선택효과',    '(성과) 종목선택효과 해설',               150),
    ('saa_운용현황',    '(운용현황) ㅇ SAA대비 운용현황 (2~3줄)',  200),
    ('시장움직임',      '(운용현황) ㅇ 월간 운용현황 — 시장 움직임 (3줄)', 420),
    ('운용현황',        '(운용현황) 운용 현황 1줄',               130),
    ('기타_주식',       '(기타) 국가/지역/종목별 주식',            90),
    ('기타_듀레이션',   '(기타) 채권 듀레이션',                    90),
    ('기타_커브',       '(기타) 채권 커브 포지션',                 80),
    ('기타_환헷지',     '(기타) 환헷지',                          100),
    ('전망_경기국면',   '2. 시장전망 ㅇ 매크로 — 경기국면',         60),
    ('전망_통화정책',   '2. 시장전망 ㅇ 매크로 — 통화정책',        300),
    ('전망_환율',       '2. 시장전망 ㅇ 매크로 — 환율',            240),
    ('전망_자산배분',   '2. 시장전망 ㅇ 자산시장 — 자산배분',      280),
    ('전망_주식',       '2. 시장전망 ㅇ 자산시장 — 주식',          280),
    ('전망_채권',       '2. 시장전망 ㅇ 자산시장 — 채권',          280),
    ('계획_자산배분',   '2. 향후 운용계획 ㅇ 자산배분',            70),
    ('계획_주식',       '2. 향후 운용계획 ㅇ 주식',               180),
    ('계획_채권',       '2. 향후 운용계획 ㅇ 채권',               180),
    ('계획_대체투자',   '2. 향후 운용계획 ㅇ 대체투자',            30),
    ('계획_환헷지',     '2. 향후 운용계획 ㅇ 환헷지',             110),
)

# ★ 운용역이 직접 쓴 문장 — **원문 그대로** 넣고 LLM 이 못 건드리게 한다.
#   매크로·전망은 외부 정보와 운용역 판단이라 LLM 이 쓰면 창작이 된다(2026-08-06 결정).
#   여기 등록된 블록은 ① 생성 결과를 덮어쓰고 ② 프롬프트에 '이미 확정된 문장'으로
#   같이 들어가 나머지 블록이 모순되지 않게 한다.
NARRATIVE_OVERRIDES: dict[str, dict[str, str]] = {
    '2026-07': {
        '전망_자산배분': (
            '중동발 유가·달러 강세와 엔화 약세에 따른 엔캐리 청산 우려는 중국의 '
            '이란 압박을 통한 호르무즈해협 개방과 미·일 공조에 따른 엔화 강세 유도로 '
            '완화될 전망입니다. 이에 글로벌 유동성 축소 및 금융시장 변동성 확대 '
            '우려가 낮아지며 위험자산 선호와 글로벌 증시 회복을 뒷받침할 것으로 '
            '예상합니다. 동시에 하이퍼스케일러의 양호한 실적과 높은 AI 투자수익률, '
            '생성형 AI 매출 성장 및 메모리 수요 강세가 확인되며 AI 설비투자 지속 '
            '가능성이 높아져 7월 반도체 조정은 저가매수 기회로 판단합니다.'
        ),
    },
}

# 코드가 쓰는 줄 — LLM 을 태우지 않는다.
_FIXED_LABELS = {
    '성과_헤드라인': '(성과) 헤드라인',
    '절대성과_환기여': '(성과) ㅇ 절대 성과 — 환기여수익률',
    'bm_초과성과': '(성과) ㅇ BM대비 초과성과',
}


def _fixed_lines(raw: dict) -> dict:
    """수치만으로 결정되는 줄 (LLM 미개입)."""
    r, contrib = raw['ret'], raw['contrib']
    fx = contrib.get('FX')
    return {
        '성과_헤드라인': (f"OCIO알아서 펀드 1개월 수익률은 {r['m1']:+.2f}% "
                     f"연초 이후 수익률은 {r['ytd']:+.2f}%"),
        '절대성과_환기여': ('환기여수익률 —' if fx is None
                       else f'환기여수익률 {fx:+.2f}%'),
        'bm_초과성과': (f"1개월 {r['m1_ex']:+.2f}%, "
                    f"연초 이후 {r['ytd_ex']:+.2f}%"),
    }


def _facts_block(raw: dict) -> str:
    """LLM 에 주는 사실 표 — 이 표 밖의 수치는 쓰지 못하게 한다."""
    r, f_, t_ = raw['ret'], raw['factors'], raw['factor_total']
    w, wp, saa = raw['weights'], raw['weights_prev'], raw['saa']
    lines = [
        '[수익률 (사실 — 이 값만 인용)]',
        f"- 펀드 1개월 {r['m1']:+.2f}% · BM {r['bm_m1']:+.2f}% · 초과 {r['m1_ex']:+.2f}%p",
        f"- 펀드 연초이후 {r['ytd']:+.2f}% · 초과 {r['ytd_ex']:+.2f}%p",
        '',
        '[자산군 기여수익률 (FX 분리 · 방법1)]',
    ]
    for k, v in (raw['contrib'] or {}).items():
        lines.append(f'- {k} {v:+.2f}%p 기여')
    lines += ['', '[성과요인분해 (자산배분/종목선택/기타)]']
    for k in FACTOR_ROWS:
        x = f_.get(k)
        if x:
            lines.append(f"- {k}: 자산배분 {x['alloc']:+.2f}%p · "
                         f"종목선택 {x['select']:+.2f}%p · 기타 {x['other']:+.2f}%p "
                         f"· 합계 {x['sum']:+.2f}%p")
    lines.append(f"- 요인별 합계: 자산배분 {t_['alloc']:+.2f}%p · "
                 f"종목선택 {t_['select']:+.2f}%p · 기타 {t_['other']:+.2f}%p")
    lines += ['', '[비중 (%) — 당월말 / SAA대비 / 전월말대비]']
    for k in WEIGHT_KEYS:
        lines.append(f'- {k} {w[k]:.1f} / {w[k] - saa[k]:+.1f} / {w[k] - wp[k]:+.1f}')
    return '\n'.join(lines)


def _prev_word_text(period: str) -> str:
    """직전 **월간** 워드 발송본 본문 — 기타 코멘트·환헷지는 승계 성격이 강하다."""
    from tools.kb_monthly_report import _latest_monthly_docx
    try:
        tpl, _ = _latest_monthly_docx(period)
    except Exception:                                     # noqa: BLE001
        return ''
    side = tpl.with_suffix(tpl.suffix + '.txt')
    if not side.exists():
        return ''
    return side.read_text(encoding='utf-8', errors='ignore')


def _seed_text(period: str) -> tuple[str, str]:
    """승인 자산군 시드 — 시장동향/전망. 미승인이면 빈 문자열."""
    from market_research.report.market_seed import load_approved_seed
    seed = load_approved_seed(period)
    if not seed:
        return '', ''
    sec = seed.get('sections') or {}

    def _join(name: str) -> str:
        d = sec.get(name) or {}
        return '\n'.join(f'- {k}: {v}' for k, v in d.items() if str(v).strip())

    return _join('market'), _join('outlook')


def _narrative_prompt(period: str, raw: dict, seed_m: str, seed_o: str,
                      prev_word: str) -> str:
    y, m = raw['end'].year, raw['end'].month
    nxt = m % 12 + 1
    ov = NARRATIVE_OVERRIDES.get(period) or {}
    labels = {k: label for k, label, _n in _BLOCKS}
    # 확정 블록은 스켈레톤에서 빼고 사실로 넘긴다 — 생성시켜 봐야 덮어쓸 뿐이고,
    # 모델이 그 내용을 모르면 나머지 블록이 반대 방향으로 흘러간다.
    skeleton = ',\n'.join(
        f'  "{k}": ""            // {label} — {n}자 내외'
        for k, label, n in _BLOCKS if k not in ov)
    fixed_txt = '\n\n'.join(f'[{labels.get(k, k)} — 운용역 확정]\n{v}'
                            for k, v in ov.items()) or '(없음)'
    return f"""너는 한국투자신탁운용 OCIO 운용역이다. KB국민은행 투자풀에 보내는
{y}년 {m}월 월간 운용보고 **워드 문서**의 서술 블록을 쓴다.

{_facts_block(raw)}

[이번 기간 승인된 시장동향 (자산군별) — 시장 움직임·전망의 근거]
{seed_m or '(없음)'}

[이번 기간 승인된 시장전망 (자산군별)]
{seed_o or '(없음)'}

[직전 월간 발송본 원문 — 문체·구조의 기준. (기타) 코멘트와 환헷지 항목은 운용
방침이 안 바뀌었으면 표현을 승계한다.]
{prev_word[:4000] or '(없음)'}

[운용역이 이미 확정한 문장 — 이 블록은 네가 쓰지 않는다. 다만 **나머지 블록이
이 판단과 모순되지 않아야** 한다(방향·톤을 여기에 맞춘다).]
{fixed_txt}

규칙
1. **수치는 위 사실 표에 있는 값만** 쓴다. 표에 없는 수치는 절대 지어내지 않는다.
   시장 레벨(KOSPI 지수, 환율 레인지)은 승인 시장동향에 있는 것만 인용한다.
2. 각 블록은 **한 항목의 본문만** 쓴다 — 'ㅇ', '-' 같은 글머리표와 라벨
   ('주식:', '채권 듀레이션:')은 붙이지 않는다. 워드에 이미 있다.
3. 자수는 지시된 값 **내외**로 맞춘다. 크게 넘기면 워드 표/여백이 깨진다.
4. 전망·계획은 {y}년 {nxt}월 기준으로 쓴다. 과거형으로 쓰지 않는다.
5. `절대성과_주식`·`절대성과_채권` 은 기여도 **원인**을 쓴다 — 기여 수치는 워드가
   문장 끝에 따로 붙이므로 본문에 다시 적지 않는다.
6. `자산배분효과`·`종목선택효과` 는 **부호와 방향이 사실 표와 일치**해야 한다.
   플러스 기여를 부정적으로 서술하지 말 것.
7. `계획_대체투자` 는 대체 비중이 0 이면 '투자계획 없음' 으로 쓴다.

아래 JSON 만 출력한다(주석 제거).
{{
{skeleton}
}}"""


def build_narrative(period: str, raw: dict, warnings: list[str],
                    *, model: str | None = None) -> dict:
    """워드 서술 블록 생성. 실패하면 빈 dict + 경고 — 표 시트는 그대로 나간다."""
    import anthropic

    from market_research.core.constants import ANTHROPIC_API_KEY, LLM_MODEL
    from market_research.core.json_utils import parse_json_response

    seed_m, seed_o = _seed_text(period)
    if not seed_m:
        warnings.append(f'{period} 자산군 시드 승인본이 없습니다 — 시장 움직임·전망 '
                        '서술의 근거가 약해집니다')
    prompt = _narrative_prompt(period, raw, seed_m, seed_o, _prev_word_text(period))

    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        resp = client.messages.create(
            model=model or LLM_MODEL,
            max_tokens=8000,
            messages=[{'role': 'user', 'content': prompt}],
        )
    except Exception as e:                                # noqa: BLE001
        warnings.append(f'워드 서술 생성 실패: {e} — 서술 시트가 비어 있습니다')
        return {}

    # ⚠ cap 잘림은 채택하지 않는다 — 잘린 JSON 이 부분 파싱되면 뒤쪽 블록이 조용히
    #   비어버린다 ([[reference_debate_token_cap]] 과 같은 함정).
    if getattr(resp, 'stop_reason', None) == 'max_tokens':
        warnings.append('워드 서술이 max_tokens 로 잘려 채택하지 않았습니다')
        return {}

    parsed = parse_json_response(resp.content[0].text) or {}
    out = {k: str(parsed.get(k) or '').strip() for k, _l, _n in _BLOCKS}
    # 운용역 확정 문장은 **생성 결과를 덮는다** (모델이 스켈레톤을 어기고 써 왔더라도).
    for k, v in (NARRATIVE_OVERRIDES.get(period) or {}).items():
        if k in out:
            out[k] = v
        else:
            warnings.append(f'override 키 "{k}" 가 워드 블록 목록에 없습니다 — 무시됨')
    empty = [k for k, v in out.items() if not v]
    if empty:
        warnings.append(f"워드 서술 {len(empty)}개 블록이 비었습니다: {', '.join(empty)}")
    over = [f'{k}({len(v)}자>{n})' for (k, _l, n), v in
            zip(_BLOCKS, (out[k] for k, _l, _n in _BLOCKS)) if len(v) > n * 1.6]
    if over:
        warnings.append(f"워드 서술 분량 초과 — {', '.join(over)}")
    return out


def _check(d: SheetData) -> None:
    """발송 전에 눈으로 봐야 하는 것만 경고로 남긴다."""
    r, tot = d.raw['ret'], d.raw['factor_total']
    gap = tot['sum'] - r['m1_ex']
    if abs(gap) > 0.05:
        d.warnings.append(
            f"요인 합계({tot['sum']:+.2f}%p)가 기준가 초과({r['m1_ex']:+.2f}%p)와 "
            f"{gap:+.2f}%p 어긋납니다 — 표3 안에서 두 기준이 섞였는지 확인하세요")
    w = d.raw['weights']
    if w['유동성'] <= 0.0:
        d.warnings.append(
            f"유동성 비중이 {w['유동성']:.1f}% 입니다 — PA 비중 합이 100%를 넘었다는 뜻으로, "
            '결제 시차일 수 있으나 발송 전 확인이 필요합니다')
    missing = [k for k in FACTOR_ROWS if k not in d.raw['factors']]
    if missing:
        d.warnings.append(f"성과요인분해에 {', '.join(missing)} 행이 없습니다 — "
                          '워드 표 행 수와 어긋납니다')


# ══════════════════════════════════════════════════════════════
# 엑셀 쓰기
# ══════════════════════════════════════════════════════════════

def _disp_len(s: str) -> int:
    """표시 폭 — 한글·전각은 2칸. openpyxl 에 autofit 이 없어 직접 잰다."""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(s))


def _pct(v: float) -> str:
    """표3 표기 — '+5.25%' (부호 포함, 소수 2자리, % 기호 포함).

    ⚠ **문자열**로 쓴다. 숫자+표시형식이면 '값만 붙여넣기' 에서 원본 숫자가 튀어나와
      발송본 표기와 달라진다 (신한 엑셀에서 확인된 함정).
    """
    return f'{v:+.2f}%'


def _num2(v: float) -> str:
    """표4 표기 — '+6.03' (% 기호 없음)."""
    return f'{v:+.2f}'


def _w1(v: float) -> str:
    """표5 비중 — '37.0' (부호 없음, 소수 1자리)."""
    return f'{v:.1f}'


def _d1(v: float) -> str:
    """표5 대비 — '+3.0' (부호 포함, 소수 1자리)."""
    return f'{v:+.1f}'


def _write_tables_sheet(ws, d: SheetData) -> None:
    """표 시트 — 워드 표3/4/5 를 격자 그대로."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style='thin', color='B0B0B0')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_f = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='44546A')
    sub_fill = PatternFill('solid', fgColor='EDEDED')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    note_f = Font(size=9, color='806000')

    r_, f_, t_ = d.raw['ret'], d.raw['factors'], d.raw['factor_total']
    y, m = d.raw['end'].year, d.raw['end'].month
    pm = d.raw['start'].year, d.raw['start'].month
    prev_lbl = f"’{(pm[0] if pm[1] > 1 else pm[0] - 1) % 100:02d}." \
               f"{(pm[1] - 1) or 12:02d}월말"
    cur_lbl = f"’{y % 100:02d}.{m:02d}월말"

    data_rows: list[int] = []
    head_rows: list[int] = []

    def put(row, col, val, *, bold=False, fill=None, align=right):
        c = ws.cell(row=row, column=col, value=val)
        c.border = box
        c.alignment = align
        if bold:
            c.font = Font(bold=True)
        if fill is not None:
            c.fill = fill
        return c

    def head(row, cols, start=1):
        for i, v in enumerate(cols):
            c = ws.cell(row=row, column=start + i, value=v)
            c.font, c.fill, c.alignment, c.border = head_f, head_fill, center, box
        head_rows.append(row)

    # ── 표3. 성과요인분해 (워드 Table 3 · 9행 × 5열) ──
    ws.cell(row=1, column=1,
            value=f'[워드 표3] 성과요인분해 {y}년 {m:02d}월').font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value='※ A1 기준 9행 × 5열. 헤더까지 워드 표와 같은 격자라 A3:E11 블록을 '
                  '그대로 복사해 붙여넣으면 됩니다. 헤더(펀드/BM/초과)는 '
                  '**기준가 기준** — 요인 합계와 정합합니다.').font = note_f

    head(3, ('구분', '펀드', 'BM', '초과 수익률', ''))
    put(4, 1, '1개월 수익률', align=center)
    put(4, 2, _num2(r_['m1']) + '%')
    put(4, 3, _num2(r_['bm_m1']) + '%')
    put(4, 4, _num2(r_['m1_ex']) + '%')
    put(4, 5, '')
    data_rows.append(4)

    put(5, 1, '초과 수익률 분해(펀드 – BM)', bold=True, fill=sub_fill, align=center)
    for c in range(2, 6):
        put(5, c, '', fill=sub_fill)

    head(6, ('요인 / 자산군', '자산배분 효과', '종목선택효과', '기타효과', '자산군별 합계'))
    for i, key in enumerate(FACTOR_ROWS, start=7):
        x = f_.get(key) or {'alloc': 0.0, 'select': 0.0, 'other': 0.0, 'sum': 0.0}
        put(i, 1, key, align=center)
        for col, k in ((2, 'alloc'), (3, 'select'), (4, 'other'), (5, 'sum')):
            put(i, col, _pct(x[k]))
        data_rows.append(i)
    put(11, 1, '요인별 합계', bold=True, fill=sub_fill, align=center)
    for col, k in ((2, 'alloc'), (3, 'select'), (4, 'other'), (5, 'sum')):
        put(11, col, _pct(t_[k]), bold=True, fill=sub_fill)
    data_rows.append(11)

    # ── 표4. 운용수익률 (워드 Table 4 · 값은 4행 C~G) ──
    ws.cell(row=13, column=1,
            value='[워드 표4] 운용수익률').font = Font(bold=True, size=12)
    ws.cell(row=14, column=1,
            value=f'※ 값 블록 = C17:G17 (7열 표의 3~7열). 설정 후* = 앵커(2022-01-03) '
                  f'기준 절대수익률입니다 — 구 발송본은 이 칸에 설정후 초과를 '
                  f'넣었습니다(2026-05 +15.44 = 초과).').font = note_f
    head(15, ('운용사', '펀드명', f'운용수익률({cur_lbl} 기준. %, %p)', '', '', '', ''))
    head(16, ('', '', '직전 1개월', 'BM대비', '연초 이후', 'BM대비', '설정 후*'))
    put(17, 1, COMPANY, align=center)
    put(17, 2, FUND_NAME, align=center)
    for col, v in ((3, r_['m1']), (4, r_['m1_ex']), (5, r_['ytd']),
                   (6, r_['ytd_ex']), (7, r_['si'])):
        put(17, col, _num2(v))
    data_rows.append(17)

    # ── 표5. 투자비중 2블록 (워드 Table 5 · 값은 3·6행 C~J) ──
    w, wp, saa = d.raw['weights'], d.raw['weights_prev'], d.raw['saa']
    ws.cell(row=19, column=1,
            value='[워드 표5] 투자 비중').font = Font(bold=True, size=12)
    ws.cell(row=20, column=1,
            value='※ 값 블록 = C23:J23(SAA 대비) · C27:J27(전월말 대비). '
                  '유동성은 잔여(100 − Σ)로 채웁니다.').font = note_f

    for base_row, label, base in ((21, '전략적자산배분(SAA) 대비', saa),
                                  (25, f'전월({prev_lbl}) 대비', wp)):
        head(base_row, ('운용사', '펀드명', f'투자 비중({cur_lbl} 기준)', '', '', '',
                        label, '', '', ''))
        head(base_row + 1, ('', '') + WEIGHT_KEYS + WEIGHT_KEYS)
        vr = base_row + 2
        put(vr, 1, COMPANY, align=center)
        put(vr, 2, FUND_NAME, align=center)
        for i, k in enumerate(WEIGHT_KEYS):
            put(vr, 3 + i, _w1(w[k]))
            put(vr, 7 + i, _d1(w[k] - base[k]))
        data_rows.append(vr)

    # 열 폭 — 표 영역만 재서(제목·※ 줄은 A 열로 흘러넘치는 텍스트라 제외)
    widths: dict[int, int] = {}
    for r in data_rows:
        for c in range(1, 11):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ''):
                widths[c] = max(widths.get(c, 0), _disp_len(v))
    for r in head_rows:
        for c in range(1, 11):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ''):
                widths[c] = max(widths.get(c, 0), _disp_len(v) // 2 + 2)
    for c in range(1, 11):
        ws.column_dimensions[get_column_letter(c)].width = \
            min(max(widths.get(c, 0) + 2, 9), 26)


def _write_narrative_sheet(ws, d: SheetData) -> None:
    """서술 시트 — 워드 위치별 1셀. 셀 하나를 통째로 복사해 문단에 붙여넣는다."""
    from openpyxl.styles import Alignment, Font, PatternFill

    head_f = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='44546A')
    fixed_fill = PatternFill('solid', fgColor='E8F0FE')     # 수치로 결정 (코드)
    manual_fill = PatternFill('solid', fgColor='E8F5E9')    # 운용역 확정 (원문)
    wrap = Alignment(wrap_text=True, vertical='top')
    ov = NARRATIVE_OVERRIDES.get(d.period) or {}

    y, m = d.raw['end'].year, d.raw['end'].month
    ws.cell(row=1, column=1, value='워드 위치').font = head_f
    ws.cell(row=1, column=2, value=f'{y}년 {m:02d}월 본문').font = head_f
    for c in (1, 2):
        ws.cell(row=1, column=c).fill = head_fill

    body_w = 110
    fixed = _fixed_lines(d.raw)
    # kind: 'calc'=코드 산출 · 'manual'=운용역 확정 원문 · 'llm'=초안
    rows: list[tuple[str, str, str]] = [
        (_FIXED_LABELS['성과_헤드라인'], fixed['성과_헤드라인'], 'calc'),
        (_FIXED_LABELS['절대성과_환기여'], fixed['절대성과_환기여'], 'calc'),
    ]
    for k, label, _n in _BLOCKS:
        rows.append((label, d.narrative.get(k, ''),
                     'manual' if k in ov else 'llm'))
        # BM대비 초과성과는 발송본에서 절대성과 바로 다음 줄이다.
        if k == '절대성과_채권':
            rows.append((_FIXED_LABELS['bm_초과성과'], fixed['bm_초과성과'], 'calc'))

    r = 2
    for label, text, kind in rows:
        a = ws.cell(row=r, column=1,
                    value=label + (' ★확정' if kind == 'manual' else ''))
        b = ws.cell(row=r, column=2, value=text)
        a.alignment = b.alignment = wrap
        if kind == 'calc':
            a.fill = b.fill = fixed_fill
        elif kind == 'manual':
            a.fill = b.fill = manual_fill
        # wrap 셀은 행 높이를 안 주면 한 줄로 접혀 보인다 (openpyxl 은 Excel 의
        # 자동맞춤을 트리거하지 못한다) → 줄 수를 재서 직접 준다.
        lines = sum(max(1, -(-_disp_len(p) // body_w))
                    for p in str(text).split('\n')) or 1
        ws.row_dimensions[r].height = min(max(lines * 15 + 6, 22), 400)
        r += 1

    ws.cell(row=r + 1, column=1,
            value='※ 파란 줄 = 수치로만 결정되는 문장(코드 산출, LLM 미개입). '
                  '초록 ★확정 줄 = 운용역이 준 원문 그대로. '
                  '나머지 흰 줄은 LLM 초안이니 손보고 쓰세요.').font = \
        Font(size=9, color='806000')
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = body_w
    ws.freeze_panes = 'A2'


def build(period: str, out_path: str | Path | None = None,
          *, narrative: bool = True) -> dict:
    """엑셀 생성 → {'path', 'warnings'}."""
    from openpyxl import Workbook

    d = collect(period)
    if narrative:
        d.narrative = build_narrative(period, d.raw, d.warnings)
    wb = Workbook()
    _write_tables_sheet(wb.active, d)
    wb.active.title = '표'
    _write_narrative_sheet(wb.create_sheet('서술'), d)

    if out_path is None:
        out_path = BASE / 'output' / OUT_NAME.format(ym=period.replace('-', ''))
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {'path': str(out_path), 'warnings': d.warnings}


if __name__ == '__main__':
    import argparse
    import sys

    sys.path.insert(0, str(BASE))
    ap = argparse.ArgumentParser()
    ap.add_argument('--month', required=True, help='YYYY-MM')
    ap.add_argument('--out', default=None)
    ap.add_argument('--no-narrative', action='store_true',
                    help='표만 — LLM 호출 생략')
    a = ap.parse_args()
    res = build(a.month, a.out, narrative=not a.no_narrative)
    print(res['path'])
    for w in res['warnings']:
        print(f'  [warn] {w}')
