# -*- coding: utf-8 -*-
"""신한라이프(2JM23) 월간 운용보고 데이터 → 엑셀 2시트 (2026-08-06 사용자 지시).

**PPT COM 치환기를 폐기하고 이 엑셀로 대체**했다. 발송본 PPT 는 그대로 쓰되,
표 값은 이 엑셀에서 **블록 복사 → PPT 표에 붙여넣기** 한다. COM 자동 치환은
PowerPoint 가 서버에 떠야 하고 수 분 걸렸으며 DRM 래핑까지 겹쳐 취약했다.

  시트1 `Comment`        — 승인 코멘트 ③ 운용성과 요약 / ⑥ 시장환경 분석·운용계획
  시트2 `자산배분현황`    — 발송본 3p 자산배분 표 + 4p '5. 주요 투자종목 현황' 표

★ 붙여넣기 규약 — **열 순서·행 순서를 발송본 표와 1:1로 맞춘다.**
  (2026-07 발송본 COM 덤프로 확정. 순서가 어긋나면 값이 엉뚱한 행에 붙는다.)

    자산배분 표(8열 14행): 자산군 | 세부자산군 | TAA(%) | 당월말 비중(%) |
                           전월말 비중(%) | 비중 변화(%p) | 성과 기여도(%p) | 비고
      → 빌더가 채우는 **자동 4열이 D~G 로 연속**이라 D:G 만 블록 복사하면 된다.
        TAA(C)·비고(H)는 수기 열이라 비워 둔다(현금 비고 상수만 채움).

    종목 표(7열): 순번 | 종목명 | 세부 자산군 | 비중(%) | 월 수익률(%) |
                  연초후 수익률 기여도(%p) | 향후 관리 방안
      → 자동 6열이 **A~F 로 연속**. G(향후 관리 방안)만 수기.

★ 값은 **문자열**로 쓴다(`39.45%`, `29.82`). 붙여넣기 모드(서식/텍스트)에 관계없이
  발송본 표기와 글자 그대로 같게 하기 위함 — 숫자+표시형식이면 '값만 붙여넣기'
  에서 원본 숫자가 튀어나온다.

★ 빌더가 만들지 않는 것 (발송본 템플릿 값 수기 유지):
    표② 운용현황(펀드수익률·변동성·BM 2종)  — 이번 엑셀 범위 밖(사용자 지시 2시트)
    TAA(%) 열 · 비고 열 · 향후 관리 방안 열   — 수기

실행:  python -m tools.shinhan_monthly_excel --month 2026-07
출력:  output/한국투자신탁운용_신한라이프_운용보고서(글로벌자산배분B형)_{YYYYMM}_데이터.xlsx
"""
from __future__ import annotations

import calendar
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

FUND = '2JM23'
FUND_NAME = '글로벌 자산배분 B형'

OUT_NAME = '한국투자신탁운용_신한라이프_운용보고서(글로벌자산배분B형)_{ym}_데이터.xlsx'

# 발송본 표 헤더 — 글자 그대로 (붙여넣기 대상과 대조용)
ALLOC_HEADER = ('자산군', '세부자산군', 'TAA(%)', '당월말 비중(%)',
                '전월말 비중(%)', '비중 변화(%p)', '성과 기여도(%p)', '비고')
SEC_HEADER = ('순번', '종목명', '세부 자산군', '비중(%)', '월 수익률(%)',
              '연초후 수익률 기여도(%p)', '향후 관리 방안')

# 현금 행 비고 — 발송본 고정 문구
CASH_NOTE = '예금 및 증거금'


@dataclass
class AllocRow:
    """자산배분 표 한 행."""
    group: str            # 주식 / 채권 / 대체 / 현금
    label: str            # 미국 성장주 …
    cur_w: float = 0.0    # 당월말 비중 %
    prev_w: float = 0.0   # 전월말 비중 %
    contrib: float = 0.0  # 성과기여도 %p

    @property
    def delta(self) -> float:
        return self.cur_w - self.prev_w


@dataclass
class SecRow:
    """주요 투자종목 표 한 행."""
    name: str
    label: str            # 세부 자산군
    weight: float         # 비중 %
    ret: float            # 월 수익률 %
    ytd_contrib: float    # 연초후 수익률 기여도 %p


@dataclass
class SheetData:
    period: str                       # 'YYYY-MM'
    start: date
    end: date
    prev_end: date
    alloc: list = field(default_factory=list)    # AllocRow
    secs: list = field(default_factory=list)     # SecRow
    comment_summary: str = ''                    # ③
    comment_outlook: str = ''                    # ⑥
    warnings: list = field(default_factory=list)


# ────────────────────────────── 기간 ──────────────────────────────

def _month_bounds(period: str) -> tuple[date, date, date]:
    """('YYYY-MM') → (당월 첫 영업일, 당월 마지막 영업일, 전월 마지막 영업일)."""
    from modules.data_loader import load_business_days_set
    y, m = int(period[:4]), int(period[5:7])
    py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
    days = sorted(load_business_days_set(
        f'{py}{pm:02d}01', f'{y}{m:02d}{calendar.monthrange(y, m)[1]:02d}'))
    cur = [d for d in days if d[:7] == f'{y}-{m:02d}']
    prev = [d for d in days if d[:7] == f'{py}-{pm:02d}']
    if not cur or not prev:
        raise ValueError(f'{period} 영업일 캘린더 부족')
    return (date.fromisoformat(cur[0]), date.fromisoformat(cur[-1]),
            date.fromisoformat(prev[-1]))


def _ymd(d: date) -> str:
    return d.strftime('%Y%m%d')


# ────────────────────────── 자산배분 표 ──────────────────────────

def _taa_weights(start: date, end: date) -> tuple[dict, dict, list]:
    """구간 PA → 표 행별 (기말비중%, 기여수익률%p) + 미매핑 경고."""
    from config.taa_classification import ppt_row_for
    from modules.data_loader import compute_single_port_pa

    pa = compute_single_port_pa(FUND, start_date=_ymd(start), end_date=_ymd(end),
                                fx_split=False)
    ss = pa.get('sec_summary')
    w: dict[str, float] = {}
    c: dict[str, float] = {}
    unmapped: list[str] = []
    if ss is None:
        return w, c, ['PA sec_summary 없음']
    for _, r in ss.iterrows():
        code = str(r.get('종목코드') or '')
        if code in ('', '유동성및기타'):
            continue          # 유동성은 잔여(100−Σ)로 계산 — 발송본 관행
        row = ppt_row_for(code)
        if row is None:
            unmapped.append(f"{code} {r.get('종목명')}")
            continue
        w[row] = w.get(row, 0.0) + float(r.get('순자산비중_끝') or 0) * 100
        c[row] = c.get(row, 0.0) + float(r.get('기여수익률') or 0) * 100
    return w, c, unmapped


def collect_alloc(start: date, end: date, prev_end: date,
                  warnings: list) -> list:
    """자산배분 표 — 당월말/전월말 비중 + 성과기여도. TAA 열은 수기라 만들지 않는다."""
    from config.taa_classification import SHINHAN_PPT_SKELETON

    cur_w, cur_c, un1 = _taa_weights(start, end)
    # 전월말 비중 = **전월 구간 PA 의 기말 비중** (2026-06 발송본으로 검증:
    # 미국 성장주 41.21 · 한국 주식 32.61 · 한국 장기채권 21.56 · 금 2.50 전부 일치).
    p_start = (prev_end.replace(day=1))
    prev_w, _, un2 = _taa_weights(p_start, prev_end)
    for u in set(un1 + un2):
        warnings.append(f'TAA 미매핑 종목 — 자산배분표에서 누락됨: {u}')

    rows = []
    for group, label in SHINHAN_PPT_SKELETON:
        if label == '현금':
            continue
        rows.append(AllocRow(group, label,
                             round(cur_w.get(label, 0.0), 2),
                             round(prev_w.get(label, 0.0), 2),
                             round(cur_c.get(label, 0.0), 2)))
    # 현금 = 잔여. 보유 합이 100 을 넘으면 음수가 된다 — 결제 시차(미지급금 25x계정)
    # 로 실제로 발생하며 원장상 정상이지만([[reference_settlement_conventions]]),
    # 고객 발송 표에 음수 현금이 찍히므로 **반드시 눈으로 확인**하게 경고한다.
    cash_cur = round(100 - sum(r.cur_w for r in rows), 2)
    cash_prev = round(100 - sum(r.prev_w for r in rows), 2)
    if cash_cur < 0 or cash_prev < 0:
        warnings.append(
            f'현금(잔여) 비중이 음수 — 당월 {cash_cur:.2f}% / 전월 {cash_prev:.2f}%. '
            f'보유 합이 100%를 넘습니다(결제 시차 가능). 발송 전 확인 필요')
    rows.append(AllocRow(
        '현금', '현금', cash_cur, cash_prev,
        round(-sum(r.contrib for r in rows) + _port_return(start, end), 2)))
    return rows


def _port_return(start: date, end: date) -> float:
    """PA 포트폴리오 개별수익률(%) — 현금 기여도 역산용."""
    from modules.data_loader import compute_single_port_pa
    pa = compute_single_port_pa(FUND, start_date=_ymd(start), end_date=_ymd(end),
                                fx_split=False)
    a = pa.get('asset_summary')
    if a is None:
        return 0.0
    hit = a[a['자산군'] == '포트폴리오']
    return float(hit['개별수익률'].iloc[0]) * 100 if len(hit) else 0.0


# ────────────────────────── 주요 종목 표 ──────────────────────────

def collect_secs(start: date, end: date, warnings: list) -> list:
    """주요 투자종목 — 비중 내림차순. 월수익률 = 당월 PA, YTD기여도 = 연초~기말 PA."""
    from config.taa_classification import classify_taa, ppt_row_for
    from modules.data_loader import compute_single_port_pa

    pa = compute_single_port_pa(FUND, start_date=_ymd(start), end_date=_ymd(end),
                                fx_split=False)
    ytd = compute_single_port_pa(FUND, start_date=f'{end.year}0101',
                                 end_date=_ymd(end), fx_split=False)
    ytd_c = {}
    yss = ytd.get('sec_summary')
    if yss is not None:
        for _, r in yss.iterrows():
            ytd_c[str(r.get('종목코드'))] = float(r.get('기여수익률') or 0) * 100

    ss = pa.get('sec_summary')
    out = []
    if ss is None:
        warnings.append('PA sec_summary 없음 — 종목표 비어 있음')
        return out
    for _, r in ss.iterrows():
        code = str(r.get('종목코드') or '')
        w = float(r.get('순자산비중_끝') or 0) * 100
        if code == '유동성및기타' or not code:
            out.append(SecRow('유동성및기타', '', round(w, 2), 0.0,
                              round(ytd_c.get('유동성및기타', 0.0), 2)))
            continue
        hit = classify_taa(code)
        out.append(SecRow(
            str(r.get('종목명') or code),
            ppt_row_for(code) or (hit[3] if hit else '(미분류)'),
            round(w, 2),
            round(float(r.get('개별수익률') or 0) * 100, 2),
            round(ytd_c.get(code, 0.0), 2)))
    out.sort(key=lambda x: -x.weight)
    # 유동성 비중은 PA 가 0 으로 주므로 잔여로 보정 (발송본 관행)
    named = [x for x in out if x.name != '유동성및기타']
    for x in out:
        if x.name == '유동성및기타':
            x.weight = round(100 - sum(n.weight for n in named), 2)
    return out


# ────────────────────────── ③⑥ 코멘트 ──────────────────────────

def collect_comments(period: str, warnings: list) -> tuple[str, str]:
    """포맷 D 승인 코멘트 → (③ 운용성과 요약, ⑥ 시장환경 분석 및 운용계획)."""
    from market_research.report.report_store import load_final
    d = load_final(period, FUND)
    if not (d and d.get('approved')):
        warnings.append(f'{period} {FUND} 코멘트 승인본 없음 — Comment 시트 비움')
        return '', ''
    text = str(d.get('final_comment') or '')
    # 포맷 D: "1. 운용성과 요약" / "2. 시장환경 분석 및 펀드운용계획"
    parts = text.split('2. 시장환경 분석 및 펀드운용계획')
    head = parts[0].replace('1. 운용성과 요약', '').strip()
    tail = parts[1].strip() if len(parts) > 1 else ''
    if not tail:
        warnings.append('코멘트에서 2번 섹션을 찾지 못함 — 포맷 D 확인 필요')
    return head, tail


# ────────────────────────── 오케스트레이터 ──────────────────────────

def collect(period: str) -> SheetData:
    """엑셀에 쓸 값 일괄 수집."""
    start, end, prev_end = _month_bounds(period)
    w: list[str] = []
    data = SheetData(period=period, start=start, end=end, prev_end=prev_end,
                     warnings=w)
    data.alloc = collect_alloc(start, end, prev_end, w)
    data.secs = collect_secs(start, end, w)
    data.comment_summary, data.comment_outlook = collect_comments(period, w)
    return data


# ══════════════════════════════════════════════════════════════
# 엑셀 쓰기
# ══════════════════════════════════════════════════════════════

def _disp_len(s: str) -> int:
    """표시 폭 — 한글·전각은 2칸으로 센다. openpyxl 에 autofit 이 없어 직접 잰다."""
    n = 0
    for ch in str(s):
        n += 2 if ord(ch) > 0x2E80 else 1
    return n


def _autofit(ws, data_rows: list[int], header_rows: list[int],
             ncols: int, min_w: int = 7, max_w: int = 44) -> None:
    """열 폭 = **표 영역만** 재서 결정. openpyxl 에 autofit 이 없어 직접 잰다.

    ⚠ 제목·※안내 줄은 A 열에 길게 들어가 옆 칸으로 흘러넘치는 텍스트다 — 폭 계산에
      넣으면 A 열이 상한까지 부풀어 표가 망가진다. 그래서 행을 명시적으로 받는다.
    ⚠ 헤더는 wrap 이라 두 줄로 접힌다 — 전체 길이를 쓰면 과하게 넓어지므로 절반만 센다.
    """
    from openpyxl.utils import get_column_letter
    widths: dict[int, int] = {}

    def _bump(col: int, w: int) -> None:
        widths[col] = max(widths.get(col, 0), w)

    for r in data_rows:
        for c in range(1, ncols + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ''):
                _bump(c, max(_disp_len(p) for p in str(v).split('\n')))
    for r in header_rows:
        for c in range(1, ncols + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ''):
                _bump(c, _disp_len(v) // 2 + 2)   # 2줄 wrap 가정
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = \
            min(max(widths.get(col, 0) + 2, min_w), max_w)


def _pct(v: float) -> str:
    """발송본 표기 — '39.45%' (부호 포함, 소수 2자리)."""
    return f'{v:.2f}%'


def _num(v: float) -> str:
    """발송본 표기 — '29.82' (퍼센트 기호 없음)."""
    return f'{v:.2f}'


def _write_comment_sheet(ws, data: SheetData) -> None:
    """Comment 시트 — 승인 코멘트를 섹션별 1셀로. PPT 텍스트 상자에 1회 붙여넣기."""
    from openpyxl.styles import Alignment, Font, PatternFill

    head = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='44546A')
    wrap = Alignment(wrap_text=True, vertical='top')

    ws['A1'] = '구분'
    ws['B1'] = f'{data.period} {FUND} ({FUND_NAME}) 운용보고 코멘트'
    for cell in (ws['A1'], ws['B1']):
        cell.font = head
        cell.fill = fill

    body_w = 130          # 코멘트 본문 폭 — 넉넉하게 (사용자 지시)
    rows = (
        ('③ 운용성과 요약', data.comment_summary),
        ('⑥ 시장환경 분석 및 펀드운용계획', data.comment_outlook),
    )
    for i, (label, text) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).alignment = wrap
        ws.cell(row=i, column=2, value=text).alignment = wrap
        # wrap 셀은 행 높이를 안 주면 한 줄로 접혀 보인다 — 줄 수를 재서 직접 준다.
        # (openpyxl 은 Excel 의 행 자동맞춤을 트리거하지 못한다)
        lines = 0
        for para in str(text).split('\n'):
            lines += max(1, -(-_disp_len(para) // body_w))   # 올림 나눗셈
        ws.row_dimensions[i].height = min(max(lines * 15 + 6, 30), 600)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = body_w
    ws.freeze_panes = 'A2'


def _write_alloc_sheet(ws, data: SheetData) -> None:
    """자산배분현황 시트 — 발송본 3p 자산배분 표 + 4p 주요 투자종목 표."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style='thin', color='B0B0B0')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_f = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='44546A')
    sub_fill = PatternFill('solid', fgColor='EDEDED')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')

    def _header(r: int, cols: tuple) -> None:
        for c, v in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font, cell.fill, cell.alignment, cell.border = \
                head_f, head_fill, center, box

    # ── 표1. 자산배분 (발송본 3p) ──
    ws.cell(row=1, column=1, value='[3p] 자산배분 현황').font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value='※ 자동 산출 = D~G 열. C(TAA)·H(비고)는 수기 열이라 비워 둡니다 '
                  '— PPT 표에는 D:G 블록만 복사해 붙여넣으세요.').font = \
        Font(size=9, color='806000')
    _header(3, ALLOC_HEADER)
    header_rows = [3]
    data_rows: list[int] = []

    r = 4
    acc = {'cur': 0.0, 'prev': 0.0, 'con': 0.0}
    tot = {'cur': 0.0, 'prev': 0.0, 'con': 0.0}

    def _emit(group: str, label: str, cur: float, prev: float, con: float,
              note: str = '', subtotal: bool = False) -> None:
        nonlocal r
        vals = (group, label, '', _pct(cur), _pct(prev), _pct(cur - prev),
                _pct(con), note)
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = box
            cell.alignment = right if 4 <= c <= 7 else center
            if subtotal:
                cell.fill = sub_fill
                cell.font = Font(bold=True)
        data_rows.append(r)
        r += 1

    prev_group = None
    for row in data.alloc:
        if prev_group is not None and row.group != prev_group:
            _emit('소계', '소계', acc['cur'], acc['prev'], acc['con'], subtotal=True)
            acc = {'cur': 0.0, 'prev': 0.0, 'con': 0.0}
        # 현금은 그룹 소계를 두지 않는다 (발송본 구조)
        note = CASH_NOTE if row.label == '현금' else ''
        _emit(row.group, row.label, row.cur_w, row.prev_w, row.contrib, note)
        for k, v in (('cur', row.cur_w), ('prev', row.prev_w), ('con', row.contrib)):
            tot[k] += v
            if row.group != '현금':
                acc[k] += v
        prev_group = row.group
    _emit('총계', '총계', tot['cur'], tot['prev'], tot['con'], subtotal=True)

    # ── 표2. 주요 투자종목 현황 (발송본 4p) ──
    r += 2
    ws.cell(row=r, column=1, value='[4p] 5. 주요 투자종목 현황').font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1,
            value='※ 자동 산출 = A~F 열. G(향후 관리 방안)는 수기 열입니다 '
                  '— PPT 표에는 A:F 블록만 복사해 붙여넣으세요.').font = \
        Font(size=9, color='806000')
    r += 1
    _header(r, SEC_HEADER)
    header_rows.append(r)
    r += 1
    for i, s in enumerate(data.secs, start=1):
        vals = (str(i), s.name, s.label, _num(s.weight), _num(s.ret),
                _num(s.ytd_contrib), '')
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = box
            cell.alignment = right if 4 <= c <= 6 else center
        data_rows.append(r)
        r += 1

    _autofit(ws, data_rows, header_rows, ncols=len(ALLOC_HEADER))


def build(period: str, out_path: str | Path | None = None) -> dict:
    """엑셀 생성 → {'path', 'warnings'}."""
    from openpyxl import Workbook

    data = collect(period)
    wb = Workbook()
    _write_comment_sheet(wb.active, data)
    wb.active.title = 'Comment'
    _write_alloc_sheet(wb.create_sheet('자산배분현황'), data)

    if out_path is None:
        out_path = (Path(__file__).resolve().parent.parent / 'output'
                    / OUT_NAME.format(ym=period.replace('-', '')))
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {'path': str(out_path), 'warnings': data.warnings}


if __name__ == '__main__':
    import argparse
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    ap = argparse.ArgumentParser()
    ap.add_argument('--month', required=True, help='YYYY-MM')
    ap.add_argument('--out', default=None)
    a = ap.parse_args()
    res = build(a.month, a.out)
    print(res['path'])
    for w in res['warnings']:
        print(f'  [warn] {w}')
