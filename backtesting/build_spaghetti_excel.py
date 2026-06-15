"""
spaghetti_paths.py 로직을 엑셀로 재현 (수식 기반).
- rt7_20260430: 일별 NAV 시계열
- rt7_AssetAlloc_20260430: 전략별 가중치 표
- 5Y(60M) 롤링, 종착 누적수익률 ≥ (1+annual)^Years-1 면 BLUE, 그 외 PINK

출력: spaghetti_paths.xlsx (수식이 살아있어 사용자가 파라미터 변경 가능)
"""
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint

ROOT = Path(__file__).parent
SRC_DATA = ROOT / "rt7_20260430"
SRC_ALLOC = ROOT / "rt7_AssetAlloc_20260430"
OUT = ROOT / "spaghetti_paths.xlsx"

TARGET = "AP"
ASSET_COLS = ["미국성장주", "국내주식", "미국국채", "미국외국채",
              "미국HY", "국내중기채", "국내장기채", "금"]
STRAT_COLS = ["BM", "SAA", "TAA", "AP"]
ALL_NUM_COLS = ASSET_COLS + STRAT_COLS  # 12개


def load_data() -> pd.DataFrame:
    df = pd.read_csv(SRC_DATA, sep="\t", thousands=",", na_values=[""])
    df.columns = [c.strip() for c in df.columns]
    df["Date"] = pd.to_datetime(df["Date"])
    for c in ALL_NUM_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["Regime"] = pd.to_numeric(df["Regime"], errors="coerce").astype("Int64")
    return df.sort_values("Date").reset_index(drop=True)


def load_alloc() -> pd.DataFrame:
    # 첫 컬럼(구분)이 공백 구분, 나머지는 탭. 수동 파싱.
    rows = []
    with open(SRC_ALLOC, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\r\n")
            if not line.strip():
                continue
            # 첫 토큰(구분)은 공백 2칸 이상으로 분리
            parts = line.split("\t")
            first = parts[0]
            # "SAA  24.40%" 처리
            head, *rest_in_first = first.split(maxsplit=1)
            if rest_in_first:
                row = [head, rest_in_first[0]] + parts[1:]
            else:
                row = [head] + parts[1:]
            rows.append(row)
    header, *body = rows
    df = pd.DataFrame(body, columns=header)
    # 비중 → float
    for c in df.columns[1:]:
        df[c] = df[c].str.rstrip("%").astype(float) / 100.0
    return df


def write_data_sheet(wb, df):
    """
    Data 시트 작성. 전략 NAV (SAA/TAA/AP) 는 가중치 × 일별수익률 누적 수식.
    - BM (K열): 외부 BM (값 유지)
    - SAA (L열): 항상 SAA 가중치 (Weights!$B$2:$I$2)
    - TAA (M열): 전일 Regime=3 → TAA(침체) 가중치 (Weights!$B$3:$I$3), 그 외 → SAA 가중치
    - AP  (N열): 전일 Regime=3 → AP(침체) 가중치 (Weights!$B$5:$I$5), 그 외 → AP 가중치 (Weights!$B$4:$I$4)
    Base row = 첫 영업일 (모든 자산 NAV 100 등장). 값 100 으로 시작.
    """
    ws = wb.create_sheet("Data")
    headers = ["Date", "Regime"] + ALL_NUM_COLS + ["YYYYMM", "is_month_end"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")

    # 전략 NAV 베이스 행 찾기 (모든 자산 NAV 첫 등장 = 모두 100)
    base_idx = df[df[ASSET_COLS].notna().all(axis=1)].index.min()
    base_excel_row = int(base_idx) + 2  # excel row

    # 자산 컬럼: C..J (8개), 전략: K=BM, L=SAA, M=TAA, N=AP
    n = len(df)
    for i, row in df.iterrows():
        excel_row = i + 2
        ws.cell(excel_row, 1, row["Date"].to_pydatetime()).number_format = "yyyy-mm-dd"
        ws.cell(excel_row, 2, int(row["Regime"]) if pd.notna(row["Regime"]) else None)
        # 개별 자산 (C..J) 및 BM (K) 은 값 그대로
        for j, c in enumerate(ASSET_COLS, start=3):
            v = row[c]
            cell = ws.cell(excel_row, j, None if pd.isna(v) else float(v))
            cell.number_format = "#,##0.00"
        # BM (K=11) 값
        v_bm = row["BM"]
        ws.cell(excel_row, 11, None if pd.isna(v_bm) else float(v_bm)).number_format = "#,##0.00"

        # SAA/TAA/AP 수식 또는 base 값
        if excel_row < base_excel_row:
            # base 이전: 비어있음
            ws.cell(excel_row, 12, None)
            ws.cell(excel_row, 13, None)
            ws.cell(excel_row, 14, None)
        elif excel_row == base_excel_row:
            # base row: 모두 100
            for j in (12, 13, 14):
                ws.cell(excel_row, j, 100.0).number_format = "#,##0.00"
        else:
            r = excel_row
            prev = r - 1
            # 옛 Excel 호환: SUMPRODUCT(arr1*arr2) 형식 (implicit intersection 회피)
            ret = f"(C{r}:J{r}/C{prev}:J{prev}-1)"
            # SAA: 항상 SAA 가중치
            f_saa = f"=L{prev}*(1+SUMPRODUCT(Weights!$B$2:$I$2*{ret}))"
            ws.cell(r, 12, f_saa).number_format = "#,##0.00"
            # TAA: 전일 Regime=3 → TAA(침체), else SAA
            f_taa = (
                f"=M{prev}*(1+IF(B{prev}=3,"
                f"SUMPRODUCT(Weights!$B$3:$I$3*{ret}),"
                f"SUMPRODUCT(Weights!$B$2:$I$2*{ret})))"
            )
            ws.cell(r, 13, f_taa).number_format = "#,##0.00"
            # AP: 전일 Regime=3 → AP(침체), else AP
            f_ap = (
                f"=N{prev}*(1+IF(B{prev}=3,"
                f"SUMPRODUCT(Weights!$B$5:$I$5*{ret}),"
                f"SUMPRODUCT(Weights!$B$4:$I$4*{ret})))"
            )
            ws.cell(r, 14, f_ap).number_format = "#,##0.00"

        # YYYYMM 수식
        ws.cell(excel_row, 15, f"=TEXT(A{excel_row},\"yyyymm\")")
        # is_month_end
        if excel_row < n + 1:
            ws.cell(excel_row, 16, f'=IF(O{excel_row}<>O{excel_row+1},1,0)')
        else:
            ws.cell(excel_row, 16, 1)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 7
    for j in range(3, 15):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.column_dimensions["O"].width = 9
    ws.column_dimensions["P"].width = 11
    ws.freeze_panes = "C2"


def write_weights_sheet(wb, alloc):
    ws = wb.create_sheet("Weights")
    ws.append(list(alloc.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")
    for _, row in alloc.iterrows():
        ws.append(list(row))
    # 비율 포맷
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            ws.cell(r, c).number_format = "0.00%"
    ws.column_dimensions["A"].width = 14
    for c in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12


def write_params_sheet(wb):
    ws = wb.create_sheet("Params")
    rows = [
        ("Target", TARGET, "분석 대상 전략 (BM/SAA/TAA/AP)"),
        ("Annual_Threshold", 0.07, "연환산 수익률 임계치"),
        ("Years", 5, "롤링 기간(년)"),
        ("Window_Months", "=B4*12", "윈도우(개월) = Years × 12"),
        ("Cumulative_Threshold", "=(1+B3)^B4-1", "누적수익률 임계 = (1+연임계)^Years - 1"),
    ]
    ws["A1"] = "파라미터"
    ws["B1"] = "값"
    ws["C1"] = "설명"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")
    for i, (k, v, desc) in enumerate(rows, start=2):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
        ws.cell(i, 3, desc)
    ws["B3"].number_format = "0.00%"
    ws["B6"].number_format = "0.00%"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50


def write_monthend_sheet(wb, df):
    ws = wb.create_sheet("MonthEnd")
    headers = ["YYYYMM", "Date"] + STRAT_COLS
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")

    # 월말 일자 추출 (Python 사전 계산: YYYYMM 라벨만)
    me_dates = df.set_index("Date").resample("ME").last().dropna(subset=[TARGET])
    me_keys = [d.strftime("%Y%m") for d in me_dates.index]

    # Data 시트 범위 (절대참조)
    n = len(df)
    rng_yyyymm = f"Data!$O$2:$O${n+1}"
    rng_date = f"Data!$A$2:$A${n+1}"
    col_map = {c: get_column_letter(3 + i) for i, c in enumerate(ALL_NUM_COLS)}
    # STRAT_COLS in Data sheet:
    strat_col_letters = {c: col_map[c] for c in STRAT_COLS}

    for i, key in enumerate(me_keys):
        r = i + 2
        ws.cell(r, 1, key)
        # XLOOKUP — _xlfn._xlws. prefix 로 spill 함수임을 명시 (Excel @ 자동삽입 회피)
        ws.cell(r, 2, f'=_xlfn._xlws.XLOOKUP($A{r},{rng_yyyymm},{rng_date},"",0,-1)').number_format = "yyyy-mm-dd"
        for j, strat in enumerate(STRAT_COLS, start=3):
            letter = strat_col_letters[strat]
            rng = f"Data!${letter}$2:${letter}${n+1}"
            ws.cell(r, j, f'=_xlfn._xlws.XLOOKUP($A{r},{rng_yyyymm},{rng},"",0,-1)').number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for j in range(3, 3 + len(STRAT_COLS)):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.freeze_panes = "C2"

    return me_keys


def write_paths_sheet(wb, me_keys, target=TARGET):
    ws = wb.create_sheet(f"Paths_{target}")
    # TARGET 컬럼 위치 (MonthEnd 시트)
    target_col_idx = 3 + STRAT_COLS.index(target)  # C=3
    target_col_letter = get_column_letter(target_col_idx)

    # 헤더: A=StartYYYYMM, B=StartDate, C..BO = m=0..60, BP=EndRet, BQ=Class
    # B는 단순 lookup, C..BO는 NAV 비율 누적, BP/BQ는 분류
    ws.cell(1, 1, "StartYYYYMM").font = Font(bold=True)
    ws.cell(1, 2, "StartDate").font = Font(bold=True)
    for m in range(0, 61):
        ws.cell(1, 3 + m, f"m={m}").font = Font(bold=True)
    end_col_idx = 3 + 61  # =64 -> BL
    cls_col_idx = end_col_idx + 1  # =65 -> BM
    ws.cell(1, end_col_idx, "EndRet(60M)").font = Font(bold=True)
    ws.cell(1, cls_col_idx, "Class").font = Font(bold=True)
    for c in range(1, cls_col_idx + 1):
        ws.cell(1, c).fill = PatternFill("solid", fgColor="DCE6F1")

    n_me = len(me_keys)
    # rolling 시작은 s=0..n_me-61 (60M 후 종착 존재)
    n_paths = n_me - 60
    me_target_rng_abs = f"MonthEnd!${target_col_letter}$2:${target_col_letter}${n_me+1}"

    for i in range(n_paths):
        r = i + 2  # excel row
        start_me_row = i + 2  # MonthEnd 시트 행번호 (1-based, header 제외)
        # A: 시작 YYYYMM
        ws.cell(r, 1, f"=MonthEnd!A{start_me_row}")
        # B: 시작일자
        ws.cell(r, 2, f"=MonthEnd!B{start_me_row}").number_format = "yyyy-mm-dd"
        # base NAV = MonthEnd!{target_col_letter}{start_me_row}
        base = f"MonthEnd!${target_col_letter}{start_me_row}"
        # m=0..60: NAV(start+m)/base - 1
        for m in range(0, 61):
            me_row = start_me_row + m
            cell = ws.cell(r, 3 + m, f"=MonthEnd!${target_col_letter}{me_row}/{base}-1")
            cell.number_format = "0.00%"
        # EndRet = m=60 셀 그대로 (편의)
        end_col_letter = get_column_letter(3 + 60)
        ws.cell(r, end_col_idx, f"={end_col_letter}{r}").number_format = "0.00%"
        # Class: IF(EndRet>=Params!$B$6, "BLUE", "PINK")
        end_letter = get_column_letter(end_col_idx)
        ws.cell(r, cls_col_idx, f'=IF({end_letter}{r}>=Params!$B$6,"BLUE","PINK")')

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 12
    for c in range(3, cls_col_idx + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9
    ws.freeze_panes = "C2"
    return n_paths


def write_spaghetti_chart_sheet(wb, me_keys, df, target=TARGET):
    """
    Spaghetti 차트: 244 paths 각각이 1개 series. Paths_AP 시트 데이터 참조.
    색상은 build 시점 정적 분류 (원본 rt7 AP NAV 기준 BLUE/PINK).
    가중치 변경 시 차트 색상은 갱신 안 됨 (빌더 재실행 필요).
    """
    ws = wb.create_sheet(f"SpaghettiChart_{target}")
    paths_ws = wb["Paths_AP"]
    n_paths = len(me_keys) - 60

    # BLUE/PINK 분류 (원본 NAV 기준)
    me = df.set_index("Date").resample("ME").last().dropna(subset=[target])
    arr = me[target].values
    th = (1.07) ** 5 - 1
    classes = ["BLUE" if arr[s + 60] / arr[s] - 1 >= th else "PINK"
               for s in range(n_paths)]
    n_b = classes.count("BLUE")
    n_p = classes.count("PINK")

    # LineChart, 244 series (from_rows=True)
    chart = LineChart()
    chart.title = f"{target} Spaghetti 5Y Paths  BLUE {n_b}/{n_paths}  PINK {n_p}/{n_paths}"
    chart.x_axis.title = "개월 (0→60, 월말)"
    chart.y_axis.title = "누적수익률"
    chart.y_axis.number_format = "0%"
    chart.height = 14
    chart.width = 26
    chart.legend = None   # 244 series legend 끔

    # 데이터: Paths_AP!C2:BK245 (244 rows × 61 cols, from_rows=True 로 각 행=1 series)
    # X 카테고리: Paths_AP!C1:BK1 (m=0..60)
    cats = Reference(paths_ws, min_col=3, max_col=63, min_row=1, max_row=1)
    data = Reference(paths_ws, min_col=3, max_col=63,
                     min_row=2, max_row=1 + n_paths)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)

    # series 별 색상 (정적)
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.marker import Marker
    # 알파 합성색 (흰 배경 가정) — 초록/핑크 계열, 한 단계 더 진하게
    # GREEN: rgba(102,204,102,0.65) -> #9CDE9C  (BLUE 카테고리 — end >= threshold)
    # PINK : rgba(247,140,140,0.85) -> #F89D9D  (PINK 카테고리 — end < threshold)
    for i, s in enumerate(chart.series):
        c = "9CDE9C" if classes[i] == "BLUE" else "F89D9D"
        s.graphicalProperties = GraphicalProperties(
            ln=LineProperties(solidFill=c, w=3175)   # 0.25pt
        )
        s.marker = Marker(symbol="none")
        s.smooth = False

    ws.add_chart(chart, "B2")
    # 정보 셀
    ws["A1"] = "Spaghetti chart"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = f"BLUE {n_b}/{n_paths} ({n_b/n_paths*100:.1f}%)  PINK {n_p}/{n_paths} ({n_p/n_paths*100:.1f}%)"
    ws.column_dimensions["A"].width = 50
    return n_paths, n_b, n_p


def write_rolling_sheet(wb, me_keys, target=TARGET):
    """
    일반 5Y 롤링수익률 시계열 (단일 라인).
    X축 = End Date, Y축 = 5Y 연환산 수익률 = (NAV_end / NAV_start)^(1/5) - 1
    시작 시점이 다른 5Y 윈도우의 종착 수익률을 시간순으로 나열.
    """
    ws = wb.create_sheet(f"Rolling_{target}")
    target_col_idx = 3 + STRAT_COLS.index(target)  # F
    target_col_letter = get_column_letter(target_col_idx)

    headers = ["EndYYYYMM", "EndDate", f"{target}_End", f"{target}_Start",
               "Cum_5Y", "Ann_5Y", "Threshold"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")

    n_me = len(me_keys)
    n_rows = n_me - 60   # 244

    for i in range(n_rows):
        r = i + 2
        end_row = r + 60      # MonthEnd 시트 행 (1-based, header excluded)
        start_row = r         # MonthEnd 시트 행
        ws.cell(r, 1, f"=MonthEnd!A{end_row}")
        ws.cell(r, 2, f"=MonthEnd!B{end_row}").number_format = "yyyy-mm-dd"
        ws.cell(r, 3, f"=MonthEnd!{target_col_letter}{end_row}").number_format = "#,##0.00"
        ws.cell(r, 4, f"=MonthEnd!{target_col_letter}{start_row}").number_format = "#,##0.00"
        ws.cell(r, 5, f"=C{r}/D{r}-1").number_format = "0.00%"
        ws.cell(r, 6, f"=(1+E{r})^(1/Params!$B$4)-1").number_format = "0.00%"
        ws.cell(r, 7, "=Params!$B$3").number_format = "0.00%"

    # LineChart: Ann_5Y + Threshold 두 라인
    chart = LineChart()
    chart.title = f"{target} 5Y 롤링 연환산 수익률"
    chart.style = 2
    chart.y_axis.title = "연환산 수익률"
    chart.x_axis.title = "End Date"
    chart.y_axis.number_format = "0.0%"
    chart.height = 10
    chart.width = 22

    data = Reference(ws, min_col=6, min_row=1, max_col=7, max_row=n_rows + 1)
    cats = Reference(ws, min_col=2, min_row=2, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # Threshold 시리즈를 점선으로
    if len(chart.series) >= 2:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        chart.series[1].graphicalProperties = GraphicalProperties(
            ln=LineProperties(prstDash="dash", solidFill="C0392B")
        )
        chart.series[0].graphicalProperties = GraphicalProperties(
            ln=LineProperties(solidFill="1F77B4", w=20000)
        )
    ws.add_chart(chart, "I2")

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 12
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "C2"

    return n_rows


def write_histogram_sheet(wb, target=TARGET):
    """
    5Y 연환산 수익률 히스토그램.
    데이터: Rolling_{target}!F2:F245 (수식 결과). COUNTIFS 로 동적 bin 카운트.
    bin: 5.0% ~ 14.0%, 0.5%p 간격 (18 bins).
    색상: lo >= 7% → green, < 7% → pink.
    """
    ws = wb.create_sheet(f"Histogram_{target}")
    ws.append(["BinLower", "BinUpper", "Label", "Freq"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")

    bins = [round(5.0 + i * 0.5, 1) for i in range(19)]  # 5.0..14.0
    n_bins = len(bins) - 1
    rolling_rng = f"Rolling_{target}!$F$2:$F$245"

    for i in range(n_bins):
        r = i + 2
        lo, hi = bins[i], bins[i + 1]
        ws.cell(r, 1, lo / 100).number_format = "0.0%"
        ws.cell(r, 2, hi / 100).number_format = "0.0%"
        ws.cell(r, 3, f"{lo:.1f}~{hi:.1f}%")
        ws.cell(r, 4,
            f'=COUNTIFS({rolling_rng},">="&A{r},{rolling_rng},"<"&B{r})')

    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = f"{target} 5Y 연환산수익률 히스토그램"
    chart.x_axis.title = "연환산수익률"
    chart.y_axis.title = "Frequency"
    chart.height = 12
    chart.width = 22
    chart.legend = None
    chart.gapWidth = 30   # 막대 사이 간격 줄임

    data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=n_bins + 1)
    cats = Reference(ws, min_col=3, min_row=2, max_row=n_bins + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # bin 별 색상 (data point 단위)
    from openpyxl.chart.shapes import GraphicalProperties
    series = chart.series[0]
    dpts = []
    for i in range(n_bins):
        lo = bins[i]
        c = "9CDE9C" if lo >= 7.0 else "F89D9D"
        dp = DataPoint(idx=i)
        dp.graphicalProperties = GraphicalProperties(solidFill=c)
        dpts.append(dp)
    series.data_points = dpts

    ws.add_chart(chart, "F2")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    return n_bins


def write_summary_sheet(wb, n_paths, target=TARGET):
    ws = wb.create_sheet("Summary")
    paths_sheet = f"Paths_{target}"
    cls_col_letter = get_column_letter(3 + 61 + 1)  # BM
    end_col_letter = get_column_letter(3 + 61)      # BL
    rng_cls = f"{paths_sheet}!${cls_col_letter}$2:${cls_col_letter}${n_paths+1}"
    rng_end = f"{paths_sheet}!${end_col_letter}$2:${end_col_letter}${n_paths+1}"

    ws["A1"] = "항목"
    ws["B1"] = "값"
    ws["C1"] = "비고"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")

    rows = [
        ("대상", f"=Params!$B$2", "전략명"),
        ("연환산 임계", f"=Params!$B$3", ""),
        ("누적 임계", f"=Params!$B$6", "(1+연임계)^Years - 1"),
        ("총 path 수", f"=COUNTA({rng_cls})", "5Y 윈도우 개수"),
        ("BLUE (>=임계) 건수", f'=COUNTIF({rng_cls},"BLUE")', ""),
        ("PINK (<임계) 건수", f'=COUNTIF({rng_cls},"PINK")', ""),
        ("BLUE 비율", f'=COUNTIF({rng_cls},"BLUE")/COUNTA({rng_cls})', ""),
        ("PINK 비율", f'=COUNTIF({rng_cls},"PINK")/COUNTA({rng_cls})', ""),
        ("EndRet 평균", f"=AVERAGE({rng_end})", "5Y 종착 누적수익률 평균"),
        ("EndRet 중앙", f"=MEDIAN({rng_end})", ""),
        ("EndRet 최소", f"=MIN({rng_end})", ""),
        ("EndRet 최대", f"=MAX({rng_end})", ""),
    ]
    for i, (k, v, d) in enumerate(rows, start=2):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
        ws.cell(i, 3, d)

    # 포맷
    ws["B3"].number_format = "0.00%"
    ws["B4"].number_format = "0.00%"
    for r in (8, 9, 10, 11, 12, 13):
        ws.cell(r, 2).number_format = "0.00%"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36


def main():
    df = load_data()
    alloc = load_alloc()
    print(f"data rows: {len(df)}, alloc rows: {len(alloc)}")

    wb = openpyxl.Workbook()
    # 기본 시트 삭제
    wb.remove(wb.active)

    write_params_sheet(wb)
    write_weights_sheet(wb, alloc)
    write_data_sheet(wb, df)
    me_keys = write_monthend_sheet(wb, df)
    print(f"month-end count: {len(me_keys)}")
    n_paths = write_paths_sheet(wb, me_keys, TARGET)
    print(f"rolling paths: {n_paths}")
    n_roll = write_rolling_sheet(wb, me_keys, TARGET)
    print(f"rolling rows: {n_roll}")
    n_sp, n_b, n_p = write_spaghetti_chart_sheet(wb, me_keys, df, TARGET)
    print(f"spaghetti chart: {n_sp} series (BLUE {n_b} / PINK {n_p})")
    n_hist = write_histogram_sheet(wb, TARGET)
    print(f"histogram bins: {n_hist}")
    write_summary_sheet(wb, n_paths, TARGET)

    wb.save(OUT)
    print(f"saved: {OUT}")


if __name__ == "__main__":
    main()
