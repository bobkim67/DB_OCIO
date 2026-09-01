# -*- coding: utf-8 -*-
"""KB국민은행 투자풀(07G07) 월간 워드 붙여넣기 엑셀 (2026-08-07).

신한(2JM23)과 같은 전환 — Word COM 치환기(`tools.kb_monthly_report.build_docx`)를
쓰지 않고, 이 엑셀에서 **블록 복사 → 발송본 워드 표에 붙여넣기** 한다.

생성 자체는 PA 를 여러 번 돌고 LLM 을 1회 태워 느리므로 여기서 다루지 않는다 —
배선·붙여넣기 격자·수치 표기 규약만 검증한다. 값은 2026-05 발송본 대조로
수동 검증했다 (요인 16셀 ±0.01%p, 절대성과 3값 정확 일치).
"""
from __future__ import annotations


def test_spec_wired_on_approve():
    from api.routers.admin_funds import _EXCEL_SPECS
    spec = _EXCEL_SPECS['07G07']
    assert spec['on'] == 'approve' and spec['kind'] == '월간'
    # 브린슨 엑셀 옵션이 붙으면 빌더 분기가 달라진다 (전용 빌더로 가야 한다)
    assert 'brinson' not in spec


def test_spec_name_matches_builder_out_name():
    """스펙 파일명 == 빌더 기본 출력명. 어긋나면 CLI 산출물과 admin 산출물이 갈린다."""
    from api.routers.admin_funds import _EXCEL_SPECS
    from tools.kb_monthly_excel import OUT_NAME
    assert _EXCEL_SPECS['07G07']['name'] == OUT_NAME


def test_excel_path_month_only():
    from api.routers.admin_funds import _excel_path
    assert _excel_path('07G07', '2026-Q2') is None
    p = _excel_path('07G07', '2026-07')
    assert p is not None and p.name.endswith('_202607.xlsx')


def test_factor_rows_match_sent_report_order():
    """워드 표3 자산군 행 순서 = 발송본 순서.

    ⚠ 엑셀은 **행 순서가 곧 정확도**다 — 워드 COM 치환기는 셀 좌표로 찍어서
      순서가 어긋나도 티가 안 났지만, 블록 복사는 값이 엉뚱한 행에 붙는다.
      (신한 엑셀 전환에서 채권 4행 순서가 뒤바뀌어 있던 것과 같은 함정)
    """
    from tools.kb_monthly_excel import FACTOR_ROWS
    assert FACTOR_ROWS == ('주식', '채권', 'FX', '유동성 및 비용')


def test_weight_keys_match_sent_report_order():
    from tools.kb_monthly_excel import WEIGHT_KEYS
    assert WEIGHT_KEYS == ('주식', '채권', '대체', '유동성')


def test_number_formats_are_strings():
    """값은 **문자열**로 쓴다.

    숫자+표시형식이면 '값만 붙여넣기' 에서 원본 숫자가 튀어나와 발송본 표기와
    달라진다 (신한 엑셀에서 확인된 함정).
    """
    from tools.kb_monthly_excel import _d1, _num2, _pct, _w1
    assert _pct(5.25) == '+5.25%' and _pct(-2.115) == '-2.12%'
    assert _num2(6.03) == '+6.03' and _num2(-7.3413) == '-7.34'
    assert _w1(37.04) == '37.0'
    assert _d1(3.0) == '+3.0' and _d1(-0.25) == '-0.2'


def test_fixed_lines_are_code_generated():
    """수치로만 결정되는 3줄은 LLM 을 안 태운다 (2JM23 성과 문장과 같은 규약)."""
    from datetime import date

    from tools.kb_monthly_excel import _fixed_lines
    raw = {'ret': {'m1': -7.3413, 'ytd': 1.7928, 'm1_ex': -3.7199,
                   'ytd_ex': 0.6419},
           'contrib': {'주식': -5.0, '채권': -0.8, 'FX': -1.52},
           'end': date(2026, 7, 31)}
    out = _fixed_lines(raw)
    assert out['성과_헤드라인'] == (
        'OCIO알아서 펀드 1개월 수익률은 -7.34% 연초 이후 수익률은 +1.79%')
    assert out['절대성과_환기여'] == '환기여수익률 -1.52%'
    assert out['bm_초과성과'] == '1개월 -3.72%, 연초 이후 +0.64%'


def test_fixed_lines_survive_missing_fx():
    """FX 행이 없어도 죽지 않는다 (FX 무보유 기간)."""
    from datetime import date

    from tools.kb_monthly_excel import _fixed_lines
    out = _fixed_lines({'ret': {'m1': 1.0, 'ytd': 2.0, 'm1_ex': 0.5,
                                'ytd_ex': 0.5},
                        'contrib': {}, 'end': date(2026, 7, 31)})
    assert out['절대성과_환기여'] == '환기여수익률 —'


def test_override_blocks_are_known_keys():
    """운용역 확정 문장(override)의 키는 워드 블록 목록에 있어야 한다.

    오타가 나면 조용히 무시돼 확정 문장이 LLM 초안으로 나간다.
    """
    from tools.kb_monthly_excel import _BLOCKS, NARRATIVE_OVERRIDES
    known = {k for k, _l, _n in _BLOCKS}
    for period, blocks in NARRATIVE_OVERRIDES.items():
        unknown = set(blocks) - known
        assert not unknown, f'{period}: {unknown}'


def test_frontend_excel_spec_matches_backend():
    """프론트 `EXCEL_SPEC` 키 == 백엔드 `_EXCEL_SPECS` 키.

    ⚠ 다운로드 버튼 노출은 **프론트의 별도 하드코딩 목록**이 결정한다. 백엔드만
      등록하면 API 는 200 을 주는데 버튼이 안 보인다 — 2026-08-07 07G07 에서 실제로
      겪었다(백엔드 excel_ready=True, 화면엔 버튼 없음).
    """
    import re
    from pathlib import Path

    from api.routers.admin_funds import _EXCEL_SPECS

    src = (Path(__file__).resolve().parents[2] / 'web' / 'src' / 'tabs'
           / 'AdminCommentWorkflowPanel.tsx').read_text(encoding='utf-8')
    block = re.search(r'const EXCEL_SPEC[^{]*\{(.*?)\n\};', src, re.S)
    assert block, 'EXCEL_SPEC 선언을 찾지 못했습니다 (프론트 구조가 바뀌었나요?)'
    front = set(re.findall(r'"([0-9A-Z]{5})"\s*:\s*\{', block.group(1)))
    assert front == set(_EXCEL_SPECS), (
        f'프론트에만: {front - set(_EXCEL_SPECS)} / '
        f'백엔드에만: {set(_EXCEL_SPECS) - front}')


def test_check_flags_factor_sum_mismatch():
    """표3 안에서 두 기준(마스터 PA / 기준가)이 섞이면 경고해야 한다.

    2026-05 발송본이 실제로 그랬다 — 헤더 초과 +4.58 vs 요인 합계 +3.76.
    """
    from datetime import date

    from tools.kb_monthly_excel import SheetData, _check
    raw = {
        'ret': {'m1': 7.39, 'bm_m1': 2.80, 'm1_ex': 4.58, 'ytd': 0.0,
                'ytd_ex': 0.0, 'si': 0.0},
        'factors': {k: {'alloc': 0.0, 'select': 0.0, 'other': 0.0, 'sum': 0.0}
                    for k in ('주식', '채권', 'FX', '유동성 및 비용')},
        'factor_total': {'alloc': -0.06, 'select': 3.11, 'other': 0.71,
                         'sum': 3.76},
        'weights': {'주식': 37.0, '채권': 62.5, '대체': 0.0, '유동성': 0.5},
        'weights_prev': {}, 'saa': {}, 'contrib': {}, 'end': date(2026, 5, 31),
    }
    d = SheetData(period='2026-05', raw=raw)
    _check(d)
    assert any('어긋납니다' in w for w in d.warnings), d.warnings


# ══════════════════════════════════════════════════════════════
# FactSheet 시트 (2026-09-01) — 워드 엑셀과 같은 파일에 시트로 얹었다
# ══════════════════════════════════════════════════════════════

def test_factsheet_asset_order_matches_form():
    """자산군 6분류 순서 = 양식 C2:H2. 순서가 어긋나면 값이 엉뚱한 칸에 붙는다."""
    from tools.kb_monthly_report import FS_ASSETS, FS_TRAIL
    assert FS_ASSETS == ('국내주식', '해외주식', '국내채권', '해외채권',
                         '대체투자', '유동성')
    assert FS_TRAIL == ('3M', '6M', '12M', '18M', '24M', '30M')


def _fake_factsheet_raw():
    from datetime import date
    a6 = {'국내주식': 21.39, '해외주식': 18.54, '국내채권': 59.28,
          '해외채권': 0.0, '대체투자': 0.0, '유동성': 0.79}
    ret_by_asset = {'국내주식': -23.23, '해외주식': -9.83, '국내채권': -2.23,
                    '해외채권': 0.0, '유동성및기타': 0.0}
    return {
        'period': '2026-07', 'start': date(2026, 7, 1), 'end': date(2026, 7, 31),
        'ret': {'m1': -7.3413, 'ytd': 1.7928, 'si': 30.1466, 'm1_ex': -3.72,
                'ytd_ex': 0.64, 'bm_m1': -3.6214, 'bm_ytd': 1.1509},
        'multi': {'3M': -3.78, '6M': 0.59, '12M': 7.71, '18M': 9.11,
                  '24M': 14.05, '30M': 24.42, 'YTD': 1.79},
        'risk': {'vol': 14.53, 'sharpe': 0.0156, 'sharpe_adj': -0.000243},
        'weights6': a6,
        'weights': {'주식': 39.93, '채권': 59.28, '대체': 0.0, '유동성': 0.79},
        # `_check` 가 워드 표3 정합도 같이 보므로 최소 키를 채워 둔다 (요인 합계 = m1_ex)
        'factors': {k: {'alloc': 0.0, 'select': 0.0, 'other': 0.0, 'sum': 0.0}
                    for k in ('주식', '채권', 'FX', '유동성 및 비용')},
        'factor_total': {'alloc': 0.0, 'select': 0.0, 'other': 0.0, 'sum': -3.72},
        'fs_1m': dict(ret_by_asset, _bm=dict(ret_by_asset)),
        'fs_ytd': dict(ret_by_asset, _bm=dict(ret_by_asset)),
        'bm_ext': {'ret': {'3M': -0.52, '6M': 0.70, '12M': 4.12, '18M': 7.08,
                           '24M': 14.41, '30M': 22.22, 'SI': 24.49},
                   'vol': 5.44, 'sharpe': -0.16, 'sharpe_adj': -0.000565},
        'duration': {'bond': 17.7, 'fund': 10.4926, 'hedge_all': 0.0},
    }


def test_factsheet_grid_matches_form_addresses():
    """양식 격자와 **셀 주소가 1:1**. 2026-07 발송본 실측 주소로 고정한다.

    A3:H3(비중) · A9:AD9(운용수익률) · A16:X16(BM) · A24:V24(동일헤지 BM).
    좌표가 밀리면 붙여넣기가 통째로 어긋나는데 눈으로는 잘 안 보인다.
    """
    from openpyxl import Workbook

    from tools.kb_monthly_excel import SheetData, _write_factsheet_sheet
    ws = Workbook().active
    _write_factsheet_sheet(ws, SheetData(period='2026-07',
                                         raw=_fake_factsheet_raw()))
    # 비중
    assert ws['B3'].value == 100.0
    assert (ws['C3'].value, ws['D3'].value, ws['E3'].value) == (21.39, 18.54, 59.28)
    assert ws['H3'].value == 0.79
    # 운용수익률 — 1M 합계/자산군, 다기간, 설정후, 연초이후, 위험, 듀레이션
    assert ws['B9'].value == -7.3413 and ws['C9'].value == -23.23
    assert (ws['I9'].value, ws['N9'].value) == (-3.78, 24.42)
    assert ws['O9'].value == 30.1466 and ws['P9'].value == 1.7928
    assert ws['W9'].value == 14.53 and ws['X9'].value == -0.000243
    assert ws['Y9'].value == 10.4926 and ws['Z9'].value == 17.7
    # BM
    assert ws['B16'].value == -3.6214 and ws['O16'].value == 24.49
    assert ws['W16'].value == 5.44 and ws['X16'].value == -0.000565
    # 동일헤지 가정 BM = 일반 BM 복사 (환헤지비율 0)
    for col in ('B', 'D', 'I', 'O', 'P', 'R'):
        assert ws[f'{col}24'].value == ws[f'{col}16'].value


def test_factsheet_writes_numbers_not_strings():
    """FactSheet 는 **숫자**로 쓴다 — 워드 표(문자열)와 규약이 반대다.

    양식이 raw float 을 담고 있어 문자열로 넣으면 붙여넣은 칸이 텍스트가 되고
    이후 계산이 깨진다.
    """
    from openpyxl import Workbook

    from tools.kb_monthly_excel import SheetData, _write_factsheet_sheet
    ws = Workbook().active
    _write_factsheet_sheet(ws, SheetData(period='2026-07',
                                         raw=_fake_factsheet_raw()))
    for addr in ('C3', 'B9', 'O9', 'Y9', 'B16', 'X16', 'B24'):
        assert isinstance(ws[addr].value, (int, float)), addr


def test_factsheet_leaves_missing_cells_blank():
    """산출 실패 칸은 0 이 아니라 **빈칸** — 0 은 '헤지 없음'처럼 유효한 값이다."""
    from openpyxl import Workbook

    from tools.kb_monthly_excel import SheetData, _write_factsheet_sheet
    raw = _fake_factsheet_raw()
    raw['bm_ext'] = {}                       # BM 시계열 로드 실패
    raw['duration'] = {'bond': None, 'fund': None, 'hedge_all': None}
    ws = Workbook().active
    _write_factsheet_sheet(ws, SheetData(period='2026-07', raw=raw))
    for addr in ('I16', 'O16', 'W16', 'X16', 'Y9', 'Z9', 'AA9'):
        assert ws[addr].value is None, addr


def test_adjusted_sharpe_branches_on_excess_sign():
    """수정샤프 = 초과>0 이면 ÷위험, 아니면 **×위험** (R module_00_Function_v3.R:1619).

    시스템 표(07G07 2026-08-31) 실측으로 고정한다 — 두 분기 모두.
    """
    from modules.data_loader import (compute_adjusted_sharpe_ratio,
                                     compute_sharpe_ratio)
    # 음수 분기 — 1M: 초과 -0.02752925 × 위험 0.114588951
    got = compute_adjusted_sharpe_ratio(0.002635904, 0.114588951, 0.030165154)
    assert abs(got - (-0.003154548)) < 1e-9
    # 양수 분기 — 18M: 초과 0.025395818 ÷ 위험 0.124306837 (일반 샤프와 동일)
    got = compute_adjusted_sharpe_ratio(0.053140615, 0.124306837, 0.027744797)
    assert abs(got - 0.204299444) < 1e-8
    assert abs(got - compute_sharpe_ratio(0.053140615, 0.124306837,
                                          0.027744797)) < 1e-12
    # 음수 분기는 일반 샤프와 **부호만 같고 크기가 다르다**
    adj = compute_adjusted_sharpe_ratio(0.027412896, 0.167904292, 0.028859891)
    plain = compute_sharpe_ratio(0.027412896, 0.167904292, 0.028859891)
    assert adj < 0 and plain < 0 and abs(adj) < abs(plain) / 10


def test_factsheet_writes_adjusted_sharpe():
    """X9·X16 에는 **수정샤프**(sharpe_adj)가 들어간다 — 일반 샤프가 아니다."""
    from openpyxl import Workbook

    from tools.kb_monthly_excel import SheetData, _write_factsheet_sheet
    raw = _fake_factsheet_raw()
    ws = Workbook().active
    _write_factsheet_sheet(ws, SheetData(period='2026-08', raw=raw))
    assert ws['X9'].value == raw['risk']['sharpe_adj']
    assert ws['X16'].value == raw['bm_ext']['sharpe_adj']
    assert ws['X9'].value != raw['risk']['sharpe']       # 일반 샤프가 새면 안 된다
